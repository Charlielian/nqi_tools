# -*- coding: utf-8 -*-
"""
数据导出模块
负责数据导出到Excel文件。

本模块同时保留两条历史导出边界：openpyxl 负责可追加、可重新打开的
工作簿编辑；xlsxwriter 负责一次性写入和流式高吞吐写入。xlsxwriter
不能在已有 xlsx 上追加 sheet，而 openpyxl 的逐单元格写入适合兼容性
和后处理、但不适合大表高速生成。所有面向调用方的导出函数都会把失败
转换为 None/False，而不是把半成品路径当成成功结果返回。

TODO: 本模块提供两套流式导出实现（stream_write_batch openpyxl 版本 与
      export_dataframe_streaming xlsxwriter 版本），但核心查询路径只使用
      export_to_excel。建议统一为 xlsxwriter 高效版本，删除冗余实现。
"""

import os
import tempfile
import logging
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from utils.config import OUTPUT_DIR
from utils.constants import (
    EXCEL_BATCH_SIZE, EXCEL_SAVE_INTERVAL, EXCEL_DEFAULT_HEADER_COLOR,
    EXCEL_DEFAULT_FONT_SIZE
)
from utils.logger import ensure_dirs
from utils.excel_styler import ExcelStyler, normalize_excel_rows

logger = logging.getLogger(__name__)


def export_to_excel(data, filename, sheet_name='Sheet1', append=False, apply_format=False):
    """导出数据到Excel文件。

    先把允许的输入统一为 DataFrame，再按“是否追加/是否格式化”选择
    引擎：已有工作簿追加必须用 openpyxl；一次性格式化走 xlsxwriter。
    调用方应把返回值当作成功凭据，失败或空数据返回 None，不会返回
    一个可能尚未完整写入的文件名。

    Args:
        data: DataFrame 或 dict with 'data' key
        filename: 文件名
        sheet_name: 工作表名称
        append: 是否追加模式（多个sheet写入同一文件）
        apply_format: 是否应用格式化（会降低性能，大数据建议设为False）

    Returns:
        str | None: 成功时返回 OUTPUT_DIR 下的完整路径，输入不支持、数据为空
            或写入异常时返回 None。
    """
    ensure_dirs()

    if isinstance(data, dict) and 'data' in data:
        df = pd.DataFrame(data['data'])
    elif isinstance(data, pd.DataFrame):
        df = data
    else:
        logger.warning("不支持的数据格式: %s", type(data))
        return None

    if df.empty:
        logger.warning("数据为空，不导出")
        return None

    filepath = os.path.join(OUTPUT_DIR, filename)

    try:
        if apply_format:
            # 一次性格式化导出（xlsxwriter，快3-5倍）
            return export_with_format(df, filename, sheet_name)
        else:
            # 快速模式：直接写入
            if append and os.path.exists(filepath):
                # openpyxl 才能在已有 xlsx 中保留工作簿并追加 sheet；xlsxwriter
                # 是只写引擎，不能读取已有文件再追加。非追加路径也明确指定
                # openpyxl，保持普通快速导出的兼容行为。
                with pd.ExcelWriter(filepath, engine='openpyxl', mode='a') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                logger.info("已追加数据到 %s", filepath)
            else:
                df.to_excel(filepath, sheet_name=sheet_name, index=False, engine='openpyxl')
                logger.info("数据已快速导出到 %s (%d行 x %d列)", filepath, len(df), len(df.columns))

            return filepath

    except Exception as e:
        logger.error("导出失败: %s", e)
        return None


def format_excel(filepath, header_color=None, font_size=None):
    """格式化Excel文件（使用ExcelStyler）。

    该函数在已有文件上重新打开并保存，因此属于 openpyxl 的后处理边界；
    它不负责生成临时文件，也不改变导出函数的成功返回值约定。

    Args:
        filepath: Excel文件路径
        header_color: 表头颜色（RGB hex，默认使用常量）
        font_size: 字体大小（默认使用常量）
    """
    if header_color is None:
        header_color = EXCEL_DEFAULT_HEADER_COLOR
    if font_size is None:
        font_size = EXCEL_DEFAULT_FONT_SIZE
    if not os.path.exists(filepath):
        logger.error("文件不存在: %s", filepath)
        return

    try:
        wb = load_workbook(filepath)
        ws = wb.active

        # 使用ExcelStyler格式化
        ExcelStyler.format_worksheet(ws)

        wb.save(filepath)
        logger.info("Excel格式化完成: %s", filepath)

    except Exception as e:
        logger.error("格式化失败: %s", e)


def export_with_format(data, filename, sheet_name='Sheet1', header_color='165DFF'):
    """导出数据到Excel并格式化（一次性写入+格式化，xlsxwriter引擎）。

    xlsxwriter 在内存中的 workbook 上同时完成写入和格式设置，不需要像
    openpyxl 后处理那样再次加载/保存。先在目标目录创建临时 xlsx，只有
    writer 正常关闭后才用 os.replace 替换目标文件；这样中断或异常不会
    覆盖一个原本可用的旧文件，也不会把未完成的 ZIP 工作簿暴露给用户。

    Returns:
        str | None: 成功完成原子替换时返回最终完整路径；任意写入、关闭或
            替换失败时清理临时文件并返回 None。
    """
    ensure_dirs()

    if isinstance(data, dict) and 'data' in data:
        df = pd.DataFrame(data['data'])
    elif isinstance(data, pd.DataFrame):
        df = data
    else:
        logger.warning("不支持的数据格式: %s", type(data))
        return None

    if df.empty:
        logger.warning("数据为空，不导出")
        return None

    filepath = os.path.join(OUTPUT_DIR, filename)
    temp_path = None
    writer = None

    try:
        # mkstemp 先在目标目录占位，关闭文件描述符并删除占位文件后，
        # 再交给 xlsxwriter 创建真正的 ZIP 工作簿。临时文件与目标文件
        # 同目录，os.replace 才能在同一文件系统内完成原子替换。
        fd, temp_path = tempfile.mkstemp(
            prefix='.nqi-', suffix='.xlsx', dir=os.path.dirname(filepath) or '.'
        )
        os.close(fd)
        os.unlink(temp_path)
        writer = pd.ExcelWriter(temp_path, engine='xlsxwriter')
        try:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            ExcelStyler.format_worksheet_xlsx(workbook, worksheet, df, header_color)
            writer.close()
            writer = None
        except Exception:
            writer.close()
            writer = None
            raise
        # writer.close() 会完成 xlsxwriter 的 ZIP 封装；只有封装完整才
        # 进入 os.replace。临时路径置 None 表示替换已成功，异常分支
        # 因此只会删除尚未交付的临时产物。
        os.replace(temp_path, filepath)
        temp_path = None
        logger.info("数据已导出到 %s (%d行 x %d列)", filepath, len(df), len(df.columns))
        return filepath

    except Exception as e:
        if writer is not None:
            try:
                writer.close()
            except Exception:
                pass
        if temp_path and os.path.exists(temp_path):
            os.unlink(temp_path)
        logger.error("导出失败: %s", e)
        return None


# ========== 流式导出支持（优化版） ==========

def create_excel_stream(filepath, sheet_name='Sheet1'):
    """创建流式写入的Excel工作簿。

    名称沿用旧接口，但当前实现返回的是 openpyxl Workbook/Worksheet，
    适用于逐行或分批填充及后续重新打开格式化；它并不是下面
    xlsxwriter ``write_row`` 高吞吐路径的工厂。filepath 仅用于日志，真正
    保存发生在 finalize 函数中。

    Returns:
        tuple: (workbook, worksheet, writer) 或失败时的 (None, None, None)
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        logger.info("[流式导出] 创建Excel文件: %s", filepath)
        return wb, ws, None
    except Exception as e:
        logger.error("[流式导出] 创建Excel文件失败: %s", e)
        return None, None, None


def stream_write_row(worksheet, row_data, row_num):
    """流式写入一行数据

    Args:
        worksheet: openpyxl worksheet 对象
        row_data: 行数据列表
        row_num: 行号（1开始）

    Returns:
        bool: 是否成功
    """
    try:
        for col, value in enumerate(row_data, start=1):
            worksheet.cell(row=row_num, column=col, value=value)
        return True
    except Exception as e:
        logger.error("[流式导出] 写入行失败: %s", e)
        return False


def stream_write_batch(worksheet, data_list, start_row, batch_size=1000):
    """批量流式写入数据（openpyxl兼容版）

    Args:
        worksheet: openpyxl worksheet 对象
        data_list: 数据列表
        start_row: 起始行号
        batch_size: 每批写入的行数

    Returns:
        int: 写入的行数
    """
    if not data_list:
        return 0

    df = pd.DataFrame(data_list)
    rows_written = 0

    for batch_start in range(0, len(df), batch_size):
        batch_end = min(batch_start + batch_size, len(df))
        batch_df = df.iloc[batch_start:batch_end]

        for row_idx, (_, row) in enumerate(batch_df.iterrows(), start=start_row + batch_start):
            for col_idx, value in enumerate(row, start=1):
                try:
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
                except Exception:
                    pass
            rows_written += 1

        logger.info("[流式导出] 已写入 %d/%d 行", rows_written, len(df))

    return rows_written


def stream_finalize_and_format(workbook, worksheet, filepath, header_color='165DFF'):
    """完成 openpyxl 流式写入并格式化。

    这是兼容旧调用方的 openpyxl 分支：先保存，再重新加载文件做样式
    后处理。返回 bool 而不是路径，因为 filepath 由调用方提供，真正需要
    判断的是保存和格式化是否都完成。
    """
    try:
        logger.info("[流式导出] 保存文件: %s", filepath)
        workbook.save(filepath)

        # 格式化
        logger.info("[流式导出] 格式化文件...")
        format_excel(filepath, header_color)
        return True
    except Exception as e:
        logger.error("[流式导出] 保存文件失败: %s", e)
        return False


def export_dataframe_streaming(df, filepath, sheet_name='Sheet1', header_color=None,
                               batch_size=None, progress_callback=None):
    """流式导出DataFrame到Excel（xlsxwriter引擎，高效版）。

    这是大表的主路径：xlsxwriter 直接用 write_row 写入分批规范化后的
    行，避免 openpyxl 的逐 cell Python 调用。batch_size 只影响内存和
    进度粒度，不改变数据单位或返回约定。与 export_with_format 相同，
    先写目标目录中的临时文件，关闭 writer 后通过 os.replace 原子交付。

    Returns:
        bool: 仅当所有批次、格式、writer.close 和原子替换均成功时为 True；
            异常时关闭 writer、删除临时文件并返回 False。
    """
    if batch_size is None:
        batch_size = EXCEL_BATCH_SIZE
    if header_color is None:
        header_color = EXCEL_DEFAULT_HEADER_COLOR

    writer = None
    temp_path = None
    try:
        logger.info("[流式导出] 开始流式导出 %d 行数据...", len(df))

        # 目标文件与临时文件同目录，才能保证 os.replace 是同一文件系统内
        # 的原子替换；在 writer.close 前绝不触碰正式路径。
        fd, temp_path = tempfile.mkstemp(prefix='.nqi-', suffix='.xlsx', dir=os.path.dirname(filepath) or '.')
        os.close(fd)
        os.unlink(temp_path)
        writer = pd.ExcelWriter(temp_path, engine='xlsxwriter')
        workbook = writer.book
        worksheet = workbook.add_worksheet(sheet_name)

        # 创建格式
        hdr_fmt = workbook.add_format({
            'bold': True,
            'font_color': '#FFFFFF',
            'bg_color': f'#{header_color}',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'text_wrap': True,
            'font_size': 11,
        })
        cell_fmt = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
        })

        # xlsxwriter 的 worksheet.write_row 接受一整行已规范化的值，适合
        # 批量写入；openpyxl 分支则保留在 stream_write_row/batch 中，供
        # 需要可编辑 Workbook 对象的旧调用方使用。
        # 写入表头
        headers = list(df.columns)
        for col_idx, col_name in enumerate(headers):
            worksheet.write(0, col_idx, col_name, hdr_fmt)

        logger.info("[流式导出] 表头写入完成，共 %d 列", len(headers))

        # 分批写入数据 - 使用write_row批量写入
        total_rows = len(df)
        for batch_start in range(0, total_rows, batch_size):
            batch_end = min(batch_start + batch_size, total_rows)
            batch_df = df.iloc[batch_start:batch_end]

            # 转成list一次性写入，比iterrows快很多
            data_rows = normalize_excel_rows(batch_df.values.tolist())
            for row_offset, row in enumerate(data_rows):
                worksheet.write_row(batch_start + row_offset + 1, 0, row, cell_fmt)

            if progress_callback:
                progress_callback(batch_end, total_rows,
                               f"正在写入... {batch_end}/{total_rows} 行")

            logger.info("[流式导出] 批次完成: %d/%d 行", batch_end, total_rows)

        # 自动列宽
        ExcelStyler.auto_adjust_column_width_xlsx(workbook, worksheet, headers, df)

        # 保存文件
        logger.info("[流式导出] 保存文件...")
        writer.close()
        os.replace(temp_path, filepath)
        temp_path = None
        logger.info("[流式导出] ✓ 文件已保存: %s", filepath)
        return True

    except Exception as e:
        if writer is not None:
            try:
                writer.close()
            except Exception:
                pass
        if temp_path and os.path.exists(temp_path):
            os.unlink(temp_path)
        logger.error("[流式导出] 导出失败: %s", e)
        return False
