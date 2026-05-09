# -*- coding: utf-8 -*-
"""
主窗口模块
提供应用程序的主界面
"""

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
import threading
import logging
import os
from datetime import datetime, timedelta
import queue

from gui.widgets import LogTextHandler, TableConfig, MultiSelectDropdown
from gui.components import SearchableCombobox, CalendarDialog, Tooltip
from gui.theme import colors, fonts, spacing
from gui.first_run import check_first_run, show_first_run_wizard
from core.auth import LoginManager
from core.query import JXCXQuery, ClusterOrderQuery
from core.export import export_with_format
from core.license import TimeMonitor, invalidate_license, verify_serial_number, write_license_from_serial
from core.license import (
    generate_machine_code, get_hw_info, verify_with_user_code,
    save_user_code, load_user_code, delete_user_code, get_user_code_info
)
from utils.logger import ensure_dirs, setup_report_logging
from utils.config import LOG_DIR, OUTPUT_DIR, EXPIRY_DATE, DEFAULT_USERNAME, DEFAULT_PASSWORD
import pandas as pd


def check_and_setup_credentials():
    """检查并设置凭证，如需首次运行引导则显示向导

    Returns:
        tuple: (needs_setup, credentials_dict)
        - needs_setup=True: 需要显示设置向导，credentials_dict为None
        - needs_setup=False: 使用默认/已保存的凭证
    """
    if check_first_run():
        success, credentials = show_first_run_wizard()
        if not success:
            return True, None
        return False, credentials
    return False, None


class NqiToolGUI:
    """NQI工具主窗口"""

    def __init__(self, root, expiry_time=None):
        self.root = root
        self.root.title("NQI工具")
        self.root.geometry("1100x800")
        self.root.minsize(800, 600)

        self.expiry_time = datetime.strptime(EXPIRY_DATE, "%Y-%m-%d") if not expiry_time else expiry_time
        self.session = None
        self.jxcx = None
        self.query_thread = None
        self.is_querying = False
        self.log_queue = queue.Queue()

        # 进度预估时间相关
        self._progress_start_time = None  # 查询开始时间
        self._progress_last_update = None  # 上次更新时间
        self._progress_last_value = 0     # 上次进度值
        self._total_expected = 0           # 预期总进度

        self._setup_logging()
        self._create_widgets()
        self._bind_events()

        # 启动时间监控（后台运行，检测时间回拨）
        self._time_monitor = TimeMonitor(interval=30, callback=self._on_time_rollback)
        self._time_monitor.start()

        self.logger.info("=" * 50)
        self.logger.info("NQI工具 GUI 启动")
        self.logger.info(f"日志文件: {self.log_file_path}")
        self.logger.info("=" * 50)

        self.load_config()

    def _setup_logging(self):
        """设置日志系统"""
        import logging
        try:
            ensure_dirs()
            log_filename = f"NqiTool_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
            self.log_file_path = os.path.join(LOG_DIR, log_filename)

            # 使用新的报表日志系统
            setup_report_logging(LOG_DIR, console=True)

            # 设置主窗口日志记录器
            self.logger = logging.getLogger('NqiTool')
            self.logger.setLevel(logging.DEBUG)

            # 添加控制台输出
            console_handler = logging.StreamHandler()
            console_handler.setLevel(logging.INFO)
            formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s',
                                        datefmt='%H:%M:%S')
            console_handler.setFormatter(formatter)
            self.logger.addHandler(console_handler)

            self.logger.info("日志系统初始化完成")
            self.logger.info(f"日志根目录: {LOG_DIR}")
        except Exception as e:
            self.logger = logging.getLogger('NqiTool')
            self.logger.setLevel(logging.DEBUG)
            self.logger.warning(f"初始化日志文件失败: {e}，日志将仅输出到界面")

    def _create_widgets(self):
        """创建界面组件 - 使用现代化设计"""
        # 顶部蓝色标题栏
        self.header = tk.Frame(self.root, bg='#165DFF', height=60)
        self.header.pack(fill=tk.X)
        self.header.pack_propagate(False)

        # 标题栏左侧 - Logo和标题
        left_frame = tk.Frame(self.header, bg='#165DFF')
        left_frame.pack(side=tk.LEFT, padx=25, pady=12)

        icon_frame = tk.Frame(left_frame, bg='#1a6ce8', width=36, height=36)
        icon_frame.pack(side=tk.LEFT, padx=(0, 12))
        icon_frame.pack_propagate(False)
        icon_label = tk.Label(icon_frame, text="📊", font=('Segoe UI Emoji', 18),
                             bg='#1a6ce8', fg='white')
        icon_label.place(relx=0.5, rely=0.5, anchor='center')

        title_frame = tk.Frame(left_frame, bg='#165DFF')
        title_frame.pack(side=tk.LEFT)

        title = tk.Label(title_frame, text="NQI工具",
                        font=('Microsoft YaHei UI', 18, 'bold'),
                        bg='#165DFF', fg='white')
        title.pack(anchor='w')

        version = tk.Label(title_frame, text="NqiTool v1.1",
                          font=('Microsoft YaHei UI', 9),
                          bg='#1a6ce8', fg='white',
                          padx=8, pady=2)
        version.pack(anchor='w', pady=(2, 0))

        # 标题栏右侧 - 状态和授权时间
        self.right_frame = tk.Frame(self.header, bg='#165DFF')
        self.right_frame.pack(side=tk.RIGHT, padx=25, pady=12)

        # 授权过期时间标签
        self.license_label = tk.Label(self.right_frame, text="",
                              font=('Microsoft YaHei UI', 9),
                              bg='#165DFF', fg='#e0e7ff')
        self.license_label.pack(side=tk.LEFT, padx=(0, 15))

        # 激活按钮
        self.activate_btn = tk.Button(self.right_frame, text="🎫 激活",
                              font=('Microsoft YaHei UI', 9),
                              bg='#22c55e', fg='white', bd=0,
                              cursor='hand2', relief='flat', padx=10, pady=4,
                              command=self._show_activate_window)
        self.activate_btn.pack(side=tk.LEFT, padx=(0, 10))

        # 状态指示器
        self.status_dot = tk.Label(self.right_frame, text="●", font=('Arial', 14),
                            bg='#165DFF', fg='#a5b4fc')
        self.status_dot.pack(side=tk.LEFT)
        self.status_text = tk.Label(self.right_frame, text="系统就绪",
                              font=('Microsoft YaHei UI', 10),
                              bg='#165DFF', fg='white')
        self.status_text.pack(side=tk.LEFT, padx=(6, 0))

        # 主内容区域
        self.main_container = tk.Frame(self.root, bg='#f9fafb')
        self.main_container.pack(fill=tk.BOTH, expand=True)

        content = tk.Frame(self.main_container, bg='#f9fafb')
        content.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        # 登录配置卡片（放在Notebook外面，顶部）
        self._build_login_card(content)

        # 创建Notebook标签页
        self.notebook = ttk.Notebook(content)
        self.notebook.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        # 标签1：数据查询
        self.jxcx_frame = tk.Frame(self.notebook, bg='#f9fafb')
        self.notebook.add(self.jxcx_frame, text=" 📊 数据查询 ")
        self._build_jxcx_content(self.jxcx_frame)

        # 标签2：聚类工单查询
        self.cluster_frame = tk.Frame(self.notebook, bg='#f9fafb')
        self.notebook.add(self.cluster_frame, text=" 📋 聚类工单查询 ")
        self._build_cluster_content(self.cluster_frame)

        # 底部：进度和日志（放在Notebook外面）
        self._build_bottom_section(content)

        # 更新授权显示
        self._update_license_display()

    def _build_card(self, parent, title=None, **kwargs):
        """创建卡片容器
        Args:
            parent: 父容器
            title: 卡片标题（可选）
            compact: 是否紧凑模式（无标题边框）
        """
        card = tk.Frame(parent, bg='white', bd=0, relief='flat')

        if title:
            header = tk.Frame(card, bg='white')
            header.pack(fill=tk.X, padx=20, pady=(16, 0))

            label = tk.Label(header, text=title,
                            font=('Microsoft YaHei UI', 13, 'bold'),
                            bg='white', fg='#374151', anchor='w')
            label.pack(fill='x')

            separator = tk.Frame(card, bg='#f3f4f6', height=1)
            separator.pack(fill=tk.X, padx=20, pady=(12, 0))

        return card

    def _build_login_card(self, parent):
        """构建登录配置卡片（一行紧凑布局）"""
        card = self._build_card(parent, "🔐 登录配置")
        card.pack(fill=tk.X, pady=(0, 10))

        body = tk.Frame(card, bg='white')
        body.pack(fill=tk.X, padx=16, pady=10)

        # 单行布局：用户名 | 密码 | 登录状态 | 按钮
        row = tk.Frame(body, bg='white')
        row.pack(fill=tk.X)

        # 用户名
        user_frame = tk.Frame(row, bg='white')
        user_frame.pack(side=tk.LEFT, padx=(0, 12))

        tk.Label(user_frame, text="用户名", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(anchor='w')
        self.username_entry = tk.Entry(user_frame, font=('Microsoft YaHei UI', 10),
                             relief='flat', bg='#f8f9fa', bd=0, width=15)
        self.username_entry.insert(0, DEFAULT_USERNAME)
        self.username_entry.pack(fill=tk.X, pady=(2, 0), ipady=4)

        # 密码
        pass_frame = tk.Frame(row, bg='white')
        pass_frame.pack(side=tk.LEFT, padx=(0, 12))

        tk.Label(pass_frame, text="密码", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(anchor='w')
        self.password_entry = tk.Entry(pass_frame, font=('Microsoft YaHei UI', 10),
                             show="●", relief='flat', bg='#f8f9fa', bd=0, width=15)
        self.password_entry.insert(0, DEFAULT_PASSWORD)
        self.password_entry.pack(fill=tk.X, pady=(2, 0), ipady=4)

        # 登录状态图标和标签
        self.login_status_icon = tk.Label(row, text="○", font=('Arial', 12, 'bold'),
                              bg='white', fg='#80868b')
        self.login_status_icon.pack(side=tk.LEFT, padx=(10, 4), pady=0)

        self.login_status_lbl = tk.Label(row, text="未登录",
                             font=('Microsoft YaHei UI', 10, 'bold'),
                             bg='white', fg='#80868b')
        self.login_status_lbl.pack(side=tk.LEFT, padx=(0, 10), pady=0)

        # 登录按钮
        self.login_btn = tk.Button(row, text="登录",
                             font=('Microsoft YaHei UI', 10, 'bold'),
                             bg='#165DFF', fg='white', bd=1,
                             relief='raised',
                             cursor='arrow', padx=20, pady=6,
                             command=self._on_login)
        self.login_btn.pack(side=tk.LEFT)

    def _build_jxcx_content(self, parent):
        """构建数据查询标签页内容"""
        params_row = tk.Frame(parent, bg='#f9fafb')
        params_row.pack(fill=tk.X, pady=(0, 10), padx=0)

        # 左侧：查询参数卡片
        self._build_query_card(params_row)

        # 右侧：提取参数卡片
        self._build_params_card(params_row)

    def _build_query_card(self, parent):
        """构建查询参数卡片（紧凑布局）"""
        card = self._build_card(parent, "🔍 查询参数")
        card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

        body = tk.Frame(card, bg='white')
        body.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)

        # ========== 数据分类（横向排列）==========
        cat_frame = tk.Frame(body, bg='white')
        cat_frame.pack(fill=tk.X, pady=(0, 8))

        tk.Label(cat_frame, text="数据分类：", font=('Microsoft YaHei UI', 9, 'bold'),
                bg='white', fg='#5f6368').pack(side=tk.LEFT, padx=(0, 6))

        self.category_vars = {}
        categories = ["干扰", "容量", "工参", "MR覆盖", "语音报表", "小区性能", "全程完好率", "语音小区"]

        for name in categories:
            var = tk.IntVar(value=0)
            self.category_vars[name] = var
            cb = tk.Checkbutton(cat_frame, text=name, variable=var,
                              font=('Microsoft YaHei UI', 9, 'bold'),
                              bg='white', fg='#202124',
                              selectcolor='#165DFF',
                              activebackground='white',
                              activeforeground='#165DFF',
                              cursor='hand2',
                              command=lambda c=name: self._on_category_changed(c))
            cb.pack(side=tk.LEFT, padx=(0, 8))

        # 添加全选/取消全选按钮
        cat_btn_frame = tk.Frame(cat_frame, bg='white')
        cat_btn_frame.pack(side=tk.RIGHT, padx=(10, 0))

        tk.Button(cat_btn_frame, text="全选",
                 font=('Microsoft YaHei UI', 8),
                 bg='#e8eaed', fg='#202124', bd=1,
                 cursor='arrow', relief='raised', padx=8, pady=2,
                 command=self._select_all_categories).pack(side=tk.LEFT, padx=(0, 3))

        tk.Button(cat_btn_frame, text="取消",
                 font=('Microsoft YaHei UI', 8),
                 bg='#e8eaed', fg='#202124', bd=1,
                 cursor='arrow', relief='raised', padx=8, pady=2,
                 command=self._deselect_all_categories).pack(side=tk.LEFT)

        # ========== 数据表选择（下拉框 + 搜索过滤）==========
        table_frame = tk.Frame(body, bg='white')
        table_frame.pack(fill=tk.X, pady=(0, 8))

        # 表头行
        table_header = tk.Frame(table_frame, bg='white')
        table_header.pack(fill=tk.X)

        tk.Label(table_header, text="选择数据表：", font=('Microsoft YaHei UI', 9, 'bold'),
                bg='white', fg='#5f6368').pack(side=tk.LEFT, padx=(0, 6))

        # 搜索框
        self.table_search_var = tk.StringVar()
        self.table_search_entry = tk.Entry(
            table_header, textvariable=self.table_search_var,
            font=('Microsoft YaHei UI', 9), width=12,
            relief='solid', bd=1
        )
        self.table_search_entry.pack(side=tk.LEFT, padx=(0, 6))
        self.table_search_entry.insert(0, "搜索报表...")
        self.table_search_entry.config(foreground='gray')

        # 搜索框事件
        self.table_search_entry.bind('<FocusIn>', self._on_search_focus_in)
        self.table_search_entry.bind('<FocusOut>', self._on_search_focus_out)
        self.table_search_entry.bind('<KeyRelease>', self._on_table_search)

        self.table_vars = {}
        TABLE_CATEGORIES = {
            '干扰': ['5G干扰小区', '4G干扰小区'],
            '容量': ['5G小区容量报表', '重要场景-天'],
            '工参': ['5G小区工参报表', '4G小区工参报表'],
            'MR覆盖': ['5GMR覆盖-小区天', '4GMR覆盖-小区天'],
            '语音报表': ['VoLTE小区监控预警', 'VONR小区监控预警', 'EPSFB小区监控预警'],
            '小区性能': ['5G小区性能KPI报表', '4G小区性能KPI报表'],
            '全程完好率': ['4G全程完好率报表', '5G全程完好率报表'],
            '语音小区': ['4G语音小区', '5G语音小区'],
            '流量热点': ['45G流量与热点评估物理站级'],
        }

        all_tables = []
        for tables in TABLE_CATEGORIES.values():
            all_tables.extend(tables)

        for name in all_tables:
            self.table_vars[name] = tk.IntVar(value=0)

        # 保存所有表格列表用于搜索
        self._all_tables = all_tables
        self._filtered_tables = all_tables.copy()

        # 使用下拉框选择数据表
        self.table_dropdown = MultiSelectDropdown(
            table_frame,
            all_tables,
            width=22,
            select_all=False
        )
        self.table_dropdown.pack(pady=(2, 0))

        # 自定义字段选择
        custom_field_frame = tk.Frame(body, bg='white')
        custom_field_frame.pack(fill=tk.X, pady=(0, 8))

        self.custom_fields_var = tk.BooleanVar(value=False)
        custom_field_cb = tk.Checkbutton(custom_field_frame, text="自定义字段",
                                        variable=self.custom_fields_var,
                                        font=('Microsoft YaHei UI', 9, 'bold'),
                                        bg='white', fg='#202124',
                                        selectcolor='#165DFF',
                                        activebackground='white',
                                        activeforeground='#165DFF',
                                        cursor='arrow',
                                        command=self._on_custom_fields_toggle)
        custom_field_cb.pack(side=tk.LEFT, padx=(0, 6))

        self.select_fields_btn = tk.Button(custom_field_frame, text="选择字段",
                                         font=('Microsoft YaHei UI', 8, 'bold'),
                                         bg='#e8eaed', fg='#202124', bd=1,
                                         cursor='arrow', relief='raised',
                                         padx=10, pady=2,
                                         state=tk.DISABLED,
                                         command=self._show_field_selector)
        self.select_fields_btn.pack(side=tk.LEFT)

        # 存储选中的字段
        self.selected_fields = {}
        self.field_configs = {}

    def _build_params_card(self, parent):
        """构建提取参数卡片（紧凑布局，右侧显示）"""
        card = self._build_card(parent, "⚙ 提取参数")
        card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))

        body = tk.Frame(card, bg='white')
        body.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)

        # 第一行：地市选择 + 快捷日期（水平排列）
        top_row = tk.Frame(body, bg='white')
        top_row.pack(fill=tk.X, pady=(0, 6))

        # 地市选择
        city_frame = tk.Frame(top_row, bg='white')
        city_frame.pack(side=tk.LEFT, padx=(0, 15))
        tk.Label(city_frame, text="地市选择", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(anchor='w')

        self.city_dropdown = MultiSelectDropdown(
            city_frame,
            MultiSelectDropdown.GD_CITIES,
            width=12,
            select_all=False
        )
        self.city_dropdown.pack(pady=(2, 0))
        self.city_dropdown.set_selected(['阳江'])

        # 快捷日期
        quick_frame = tk.Frame(top_row, bg='white')
        quick_frame.pack(side=tk.LEFT, padx=(0, 15))
        tk.Label(quick_frame, text="快捷日期", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(anchor='w')

        quick_inner = tk.Frame(quick_frame, bg='white')
        quick_inner.pack(pady=(2, 0))

        self.quick_date_btns = {}
        for text, days in [("昨天", 1), ("近7天", 7), ("近30天", 30)]:
            btn = tk.Button(quick_inner, text=text, font=('Microsoft YaHei UI', 8, 'bold'),
                           bg='#e8eaed', fg='#202124', bd=1, padx=10, pady=2,
                           cursor='arrow', relief='raised',
                           command=lambda d=days: self.set_quick_date(d))
            btn.pack(side=tk.LEFT, padx=(0, 3))
            self.quick_date_btns[days] = btn

        # 第二行：日期范围（单独一行）
        date_row = tk.Frame(body, bg='white')
        date_row.pack(fill=tk.X, pady=(0, 6))

        # 日期范围
        date_frame = tk.Frame(date_row, bg='white')
        date_frame.pack(side=tk.LEFT, padx=(0, 15))
        tk.Label(date_frame, text="日期范围", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(anchor='w')

        date_inner = tk.Frame(date_frame, bg='white')
        date_inner.pack(pady=(2, 0))

        self.start_year_var = tk.IntVar(value=datetime.now().year)
        self.start_month_var = tk.IntVar(value=datetime.now().month)
        self.start_day_var = tk.IntVar(value=1)

        yesterday = datetime.now() - timedelta(days=1)
        self.end_year_var = tk.IntVar(value=yesterday.year)
        self.end_month_var = tk.IntVar(value=yesterday.month)
        self.end_day_var = tk.IntVar(value=yesterday.day)

        start_frame = tk.Frame(date_inner, bg='white')
        start_frame.pack(side=tk.LEFT)

        current_year = datetime.now().year
        ttk.Combobox(start_frame, textvariable=self.start_year_var,
                   values=list(range(2020, current_year + 1)),
                   width=4, state="readonly").pack(side=tk.LEFT)
        tk.Label(start_frame, text="-", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(side=tk.LEFT, padx=1)
        ttk.Combobox(start_frame, textvariable=self.start_month_var,
                   values=list(range(1, 13)),
                   width=2, state="readonly").pack(side=tk.LEFT)
        tk.Label(start_frame, text="-", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(side=tk.LEFT, padx=1)
        ttk.Combobox(start_frame, textvariable=self.start_day_var,
                   values=list(range(1, 32)),
                   width=2, state="readonly").pack(side=tk.LEFT)

        # 开始日期日历按钮
        start_cal_btn = tk.Button(start_frame, text="📅",
                                 font=('Arial', 10), bg='white', fg='#165DFF',
                                 bd=0, cursor='hand2', relief='flat',
                                 command=lambda: self._show_calendar('start'))
        start_cal_btn.pack(side=tk.LEFT, padx=(4, 0))
        Tooltip(start_cal_btn, "点击选择开始日期")

        tk.Label(date_inner, text=" 至 ", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(side=tk.LEFT, padx=3)

        end_frame = tk.Frame(date_inner, bg='white')
        end_frame.pack(side=tk.LEFT)

        ttk.Combobox(end_frame, textvariable=self.end_year_var,
                   values=list(range(2020, current_year + 1)),
                   width=4, state="readonly").pack(side=tk.LEFT)
        tk.Label(end_frame, text="-", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(side=tk.LEFT, padx=1)
        ttk.Combobox(end_frame, textvariable=self.end_month_var,
                   values=list(range(1, 13)),
                   width=2, state="readonly").pack(side=tk.LEFT)
        tk.Label(end_frame, text="-", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(side=tk.LEFT, padx=1)
        ttk.Combobox(end_frame, textvariable=self.end_day_var,
                   values=list(range(1, 32)),
                   width=2, state="readonly").pack(side=tk.LEFT)

        # 结束日期日历按钮
        end_cal_btn = tk.Button(end_frame, text="📅",
                               font=('Arial', 10), bg='white', fg='#165DFF',
                               bd=0, cursor='hand2', relief='flat',
                               command=lambda: self._show_calendar('end'))
        end_cal_btn.pack(side=tk.LEFT, padx=(4, 0))
        Tooltip(end_cal_btn, "点击选择结束日期")

        # 第三行：字段获取方式
        field_mode_row = tk.Frame(body, bg='white')
        field_mode_row.pack(fill=tk.X, pady=(0, 6))

        field_mode_frame = tk.Frame(field_mode_row, bg='white')
        field_mode_frame.pack(side=tk.LEFT, padx=(0, 15))
        tk.Label(field_mode_frame, text="字段获取方式", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(anchor='w')

        field_mode_inner = tk.Frame(field_mode_frame, bg='white')
        field_mode_inner.pack(pady=(2, 0))

        self.field_mode_var = tk.StringVar(value='hardcode')
        hardcode_rb = tk.Radiobutton(field_mode_inner, text="硬编码",
                                     variable=self.field_mode_var, value='hardcode',
                                     font=('Microsoft YaHei UI', 8, 'bold'),
                                     bg='white', fg='#202124',
                                     activebackground='white',
                                     cursor='hand2',
                                     command=self._on_field_mode_changed)
        hardcode_rb.pack(side=tk.LEFT, padx=(0, 10))

        dynamic_rb = tk.Radiobutton(field_mode_inner, text="动态获取",
                                   variable=self.field_mode_var, value='dynamic',
                                   font=('Microsoft YaHei UI', 8, 'bold'),
                                   bg='white', fg='#202124',
                                   activebackground='white',
                                   cursor='hand2',
                                   command=self._on_field_mode_changed)
        dynamic_rb.pack(side=tk.LEFT)

        # 第三行：多日模式选项（在日期范围下方，按钮上方）
        mode_row = tk.Frame(body, bg='white')
        mode_row.pack(fill=tk.X, pady=(0, 6))

        mode_frame = tk.Frame(mode_row, bg='white')
        mode_frame.pack(side=tk.LEFT, padx=(0, 0))

        self.multi_day_var = tk.BooleanVar(value=False)
        multi_day_cb = tk.Checkbutton(mode_frame, text="按日查询",
                                      variable=self.multi_day_var,
                                      font=('Microsoft YaHei UI', 8),
                                      bg='white', fg='#202124',
                                      selectcolor='#e8f0fe',
                                      activebackground='white',
                                      command=self._on_multi_day_toggle)
        multi_day_cb.pack(side=tk.LEFT)

        self.multi_day_per_sheet_var = tk.BooleanVar(value=False)
        multi_day_per_sheet_cb = tk.Checkbutton(mode_frame, text="按日分Sheet",
                                               variable=self.multi_day_per_sheet_var,
                                               font=('Microsoft YaHei UI', 8),
                                               bg='white', fg='#202124',
                                               selectcolor='#e8f0fe',
                                               activebackground='white',
                                               state=tk.DISABLED,
                                               command=self._on_multi_day_per_sheet_toggle)
        self.multi_day_per_sheet_cb = multi_day_per_sheet_cb
        multi_day_per_sheet_cb.pack(side=tk.LEFT, padx=(6, 0))

        # 第四行：操作按钮
        btn_row = tk.Frame(body, bg='white')
        btn_row.pack(fill=tk.X, pady=(4, 0))

        self.extract_btn = tk.Button(btn_row, text="▶ 开始提取",
                               font=('Microsoft YaHei UI', 10, 'bold'),
                               bg='#165DFF', fg='white', bd=1,
                               cursor='arrow', relief='raised', padx=22, pady=5,
                               state=tk.DISABLED, command=self._on_query)
        self.extract_btn.pack(side=tk.LEFT)

        self.stop_btn = tk.Button(btn_row, text="⏹ 停止",
                            font=('Microsoft YaHei UI', 9),
                            bg='#dc3545', fg='white', bd=1,
                            cursor='arrow', relief='raised', padx=14, pady=5,
                            state=tk.DISABLED, command=self._on_stop)
        self.stop_btn.pack(side=tk.LEFT, padx=(8, 0))

        tk.Button(btn_row, text="📁 打开目录",
                 font=('Microsoft YaHei UI', 9),
                 bg='#f0f2f5', fg='#202124', bd=1,
                 cursor='arrow', relief='raised', padx=12, pady=5,
                 command=self.open_output_dir).pack(side=tk.RIGHT)

    def _build_cluster_content(self, parent):
        """构建聚类工单查询标签页内容"""
        # 主容器
        main_frame = tk.Frame(parent, bg='#f9fafb')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)

        # ========== 顶部查询参数区域 ==========
        params_card = self._build_card(main_frame, "🔍 聚类工单查询参数")
        params_card.pack(fill=tk.X, pady=(0, 10))

        body = tk.Frame(params_card, bg='white')
        body.pack(fill=tk.X, padx=16, pady=12)

        # 第一行：时间类型 + 日期范围
        row1 = tk.Frame(body, bg='white')
        row1.pack(fill=tk.X, pady=(0, 8))
        row1.pack_propagate(False)

        # 时间类型
        time_type_frame = tk.Frame(row1, bg='white')
        time_type_frame.pack(side=tk.LEFT, padx=(0, 15))
        tk.Label(time_type_frame, text="时间类型", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(anchor='w')

        self.cluster_time_type_var = tk.StringVar(value='问题生成时间')
        time_type_menu = ttk.Combobox(time_type_frame,
                                       textvariable=self.cluster_time_type_var,
                                       values=['问题生成时间', '派发时间', '确认时间'],
                                       width=12, state='readonly')
        time_type_menu.pack(pady=(2, 0))

        # 日期范围
        date_frame = tk.Frame(row1, bg='white')
        date_frame.pack(side=tk.LEFT, padx=(0, 15))
        tk.Label(date_frame, text="日期范围", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(anchor='w')

        date_inner = tk.Frame(date_frame, bg='white')
        date_inner.pack(pady=(2, 0))

        yesterday = datetime.now() - timedelta(days=1)
        self.cluster_start_var = tk.StringVar(value=yesterday.strftime('%Y-%m-%d'))
        self.cluster_end_var = tk.StringVar(value=yesterday.strftime('%Y-%m-%d'))

        start_entry = tk.Entry(date_inner, textvariable=self.cluster_start_var,
                              font=('Microsoft YaHei UI', 9), width=10)
        start_entry.pack(side=tk.LEFT)

        tk.Label(date_inner, text=" 至 ", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(side=tk.LEFT, padx=3)

        end_entry = tk.Entry(date_inner, textvariable=self.cluster_end_var,
                            font=('Microsoft YaHei UI', 9), width=10)
        end_entry.pack(side=tk.LEFT)

        # 第二行：地市 + 责任网格
        row2 = tk.Frame(body, bg='white')
        row2.pack(fill=tk.X, pady=(0, 8))
        row2.pack_propagate(False)

        # 地市选择
        city_frame = tk.Frame(row2, bg='white')
        city_frame.pack(side=tk.LEFT, padx=(0, 15))
        tk.Label(city_frame, text="地市", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(anchor='w')

        # 地市编码映射
        self.CITY_CODE_MAP = {
            '广州': '860199', '深圳': '860755', '东莞': '860769',
            '佛山': '860757', '中山': '860760', '珠海': '860756',
            '惠州': '860752', '江门': '860750', '肇庆': '860758',
            '汕头': '860754', '汕尾': '860660', '潮州': '860761',
            '揭阳': '860663', '云浮': '860766', '湛江': '860759',
            '茂名': '860668', '阳江': '860662', '韶关': '860751',
            '清远': '860762', '梅州': '860753', '河源': '860670',
            '广东': ''  # 全网
        }

        self.cluster_city_var = tk.StringVar(value='阳江')
        self.cluster_city_menu = ttk.Combobox(city_frame,
                                              textvariable=self.cluster_city_var,
                                              values=list(self.CITY_CODE_MAP.keys()),
                                              width=10, state='readonly')
        self.cluster_city_menu.pack(pady=(2, 0))
        self.cluster_city_menu.bind('<<ComboboxSelected>>', self._on_cluster_city_changed)

        # 责任网格（多选）
        grid_frame = tk.Frame(row2, bg='white')
        grid_frame.pack(side=tk.LEFT, padx=(0, 15))
        tk.Label(grid_frame, text="责任网格", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(anchor='w')

        self.cluster_grid_var = tk.StringVar(value='')
        self.cluster_grid_entry = tk.Entry(grid_frame,
                                          textvariable=self.cluster_grid_var,
                                          font=('Microsoft YaHei UI', 9), width=18)
        self.cluster_grid_entry.pack(pady=(2, 0))

        tk.Label(grid_frame, text="（留空表示全部）",
                font=('Microsoft YaHei UI', 7), bg='white', fg='#9ca3af').pack(anchor='w')

        # 第三行：问题类型
        row3 = tk.Frame(body, bg='white')
        row3.pack(fill=tk.X, pady=(0, 8))

        # 问题类型选择
        type_frame = tk.Frame(row3, bg='white')
        type_frame.pack(side=tk.LEFT, padx=(0, 15))
        tk.Label(type_frame, text="问题类型", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(anchor='w')

        self.cluster_problem_type_var = tk.StringVar(value='')
        self.cluster_problem_type_entry = tk.Entry(type_frame,
                                                   textvariable=self.cluster_problem_type_var,
                                                   font=('Microsoft YaHei UI', 9), width=18)
        self.cluster_problem_type_entry.pack(pady=(2, 0))

        # 加载问题标签按钮
        self.cluster_load_labels_btn = tk.Button(row3, text="加载问题标签",
                                   font=('Microsoft YaHei UI', 8),
                                   bg='#e8eaed', fg='#202124', bd=1,
                                   cursor='arrow', relief='raised', padx=8, pady=2,
                                   command=self._load_cluster_problem_labels)
        self.cluster_load_labels_btn.pack(side=tk.LEFT, padx=(5, 0))

        # 问题状态
        status_frame = tk.Frame(row3, bg='white')
        status_frame.pack(side=tk.LEFT, padx=(0, 15))
        tk.Label(status_frame, text="工单状态", font=('Microsoft YaHei UI', 8),
                bg='white', fg='#5f6368').pack(anchor='w')

        self.cluster_status_var = tk.StringVar(value='')
        self.cluster_status_menu = ttk.Combobox(status_frame,
                                               textvariable=self.cluster_status_var,
                                               values=['', '待处理', '处理中', '已解决', '已关闭'],
                                               width=10, state='readonly')
        self.cluster_status_menu.pack(pady=(2, 0))

        # 第四行：操作按钮
        row4 = tk.Frame(body, bg='white')
        row4.pack(fill=tk.X, pady=(4, 0))

        self.cluster_query_btn = tk.Button(row4, text="🔍 查询",
                                 font=('Microsoft YaHei UI', 10, 'bold'),
                                 bg='#165DFF', fg='white', bd=1,
                                 cursor='arrow', relief='raised', padx=22, pady=5,
                                 command=self._on_cluster_query)
        self.cluster_query_btn.pack(side=tk.LEFT)

        self.cluster_export_btn = tk.Button(row4, text="📥 导出",
                                  font=('Microsoft YaHei UI', 10, 'bold'),
                                  bg='#22c55e', fg='white', bd=1,
                                  cursor='arrow', relief='raised', padx=22, pady=5,
                                  state=tk.DISABLED,
                                  command=self._on_cluster_export)
        self.cluster_export_btn.pack(side=tk.LEFT, padx=(8, 0))

        self.cluster_stop_btn = tk.Button(row4, text="⏹ 停止",
                                font=('Microsoft YaHei UI', 9),
                                bg='#dc3545', fg='white', bd=1,
                                cursor='arrow', relief='raised', padx=14, pady=5,
                                state=tk.DISABLED,
                                command=self._on_cluster_stop)
        self.cluster_stop_btn.pack(side=tk.LEFT, padx=(8, 0))

        # ========== 中部结果显示区域 ==========
        result_card = self._build_card(main_frame, "📊 查询结果")
        result_card.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        result_body = tk.Frame(result_card, bg='white')
        result_body.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)

        # 创建Treeview表格
        table_frame = tk.Frame(result_body, bg='white')
        table_frame.pack(fill=tk.BOTH, expand=True)

        # 滚动条
        y_scroll = tk.Scrollbar(table_frame, orient=tk.VERTICAL)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        x_scroll = tk.Scrollbar(table_frame, orient=tk.HORIZONTAL)
        x_scroll.pack(side=tk.BOTTOM, fill=tk.X)

        # 定义列
        self.cluster_columns = [
            '序号', '聚类工单序号', '问题小区', '问题小区名', '地市', '责任网格',
            '派发时间', '问题类型', '方案类型', '方案描述', '方案确认人',
            '评估状态', '当前状态', '问题原因', '数据来源'
        ]

        self.cluster_tree = ttk.Treeview(table_frame,
                                        columns=self.cluster_columns,
                                        show='tree headings',
                                        yscrollcommand=y_scroll.set,
                                        xscrollcommand=x_scroll.set,
                                        height=12)

        y_scroll.config(command=self.cluster_tree.yview)
        x_scroll.config(command=self.cluster_tree.xview)

        # 设置列宽
        for col in self.cluster_columns:
            self.cluster_tree.heading(col, text=col, anchor='w')
            self.cluster_tree.column(col, width=120, anchor='w')

        self.cluster_tree.pack(fill=tk.BOTH, expand=True)

        # 分页控制
        pager_frame = tk.Frame(result_body, bg='white')
        pager_frame.pack(fill=tk.X, pady=(8, 0))

        self.cluster_page_label = tk.Label(pager_frame, text="第 1 / 1 页，共 0 条",
                                          font=('Microsoft YaHei UI', 9),
                                          bg='white', fg='#5f6368')
        self.cluster_page_label.pack(side=tk.LEFT)

        tk.Button(pager_frame, text="◀ 上一页",
                 font=('Microsoft YaHei UI', 8),
                 bg='#e8eaed', fg='#202124', bd=1,
                 cursor='arrow', relief='raised', padx=8, pady=2,
                 state=tk.DISABLED,
                 command=self._on_cluster_prev_page).pack(side=tk.LEFT, padx=(10, 3))

        tk.Button(pager_frame, text="下一页 ▶",
                 font=('Microsoft YaHei UI', 8),
                 bg='#e8eaed', fg='#202124', bd=1,
                 cursor='arrow', relief='raised', padx=8, pady=2,
                 state=tk.DISABLED,
                 command=self._on_cluster_next_page).pack(side=tk.LEFT)

        # 存储查询结果
        self.cluster_all_data = []  # 所有数据（翻页用）
        self.cluster_current_page = 1
        self.cluster_total_pages = 1
        self.cluster_total_count = 0
        self.cluster_is_querying = False
        self.cluster_problem_labels = []

    def _build_bottom_section(self, parent):
        """构建底部日志区域"""
        # 直接使用 Frame 作为容器，填满所有剩余空间
        bottom = tk.Frame(parent, bg='white')
        bottom.pack(fill=tk.BOTH, expand=True, pady=(0, 0))

        progress_area = tk.Frame(bottom, bg='white')
        progress_area.pack(fill=tk.X, padx=16, pady=(10, 6))

        progress_info = tk.Frame(progress_area, bg='white')
        progress_info.pack(fill=tk.X)

        self.progress_lbl_pct = tk.Label(progress_info, text="进度: 0%",
                          font=('Microsoft YaHei UI', 10, 'bold'),
                          bg='white', fg='#165DFF')
        self.progress_lbl_pct.pack(side=tk.LEFT)

        self.progress_lbl_detail = tk.Label(progress_info, text="就绪",
                             font=('Microsoft YaHei UI', 9),
                             bg='white', fg='#5f6368')
        self.progress_lbl_detail.pack(side=tk.RIGHT)

        # 预估剩余时间标签
        self.progress_lbl_eta = tk.Label(progress_info, text="",
                             font=('Microsoft YaHei UI', 9),
                             bg='white', fg='#6b7280')
        self.progress_lbl_eta.pack(side=tk.RIGHT, padx=(0, 10))

        # 圆角进度条
        self.progress_canvas = tk.Canvas(progress_area, height=8, bg='white', highlightthickness=0)
        self.progress_canvas.pack(fill=tk.X, pady=(6, 0))
        self.progress_bar = self.progress_canvas.create_rectangle(0, 0, 0, 8, fill='#165DFF', outline='')
        self.progress_bg = self.progress_canvas.create_rectangle(0, 0, 1000, 8, fill='#f0f2f5', outline='')
        self.progress_canvas_width = 0  # 动态获取

        # 日志输出
        log_area = tk.Frame(bottom, bg='white')
        log_area.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 16))

        self.log_text = scrolledtext.ScrolledText(log_area, height=15,
                                                  font=('Consolas', 9),
                                                  state='disabled',
                                                  bg='#f8f9fa',
                                                  relief='flat',
                                                  bd=1)
        self.log_text.pack(fill=tk.BOTH, expand=True)

        # 添加日志处理器
        handler = LogTextHandler(self.log_text)
        handler.setLevel(logging.INFO)
        self.logger.addHandler(handler)

    def _on_cluster_city_changed(self, event=None):
        """地市选择改变时触发"""
        city_name = self.cluster_city_var.get()
        self.logger.info("[聚类工单] 选择地市: %s", city_name)

    def _load_cluster_problem_labels(self):
        """加载问题标签列表"""
        if not self.session:
            messagebox.showwarning("提示", "请先登录后再操作")
            return

        start_date = self.cluster_start_var.get()
        end_date = self.cluster_end_var.get()

        if not start_date or not end_date:
            messagebox.showwarning("提示", "请选择日期范围")
            return

        self.logger.info("[聚类工单] 正在加载问题标签...")
        self.cluster_load_labels_btn.config(state=tk.DISABLED)

        def do_load():
            try:
                from core.query import ClusterOrderQuery
                cluster_query = ClusterOrderQuery(self.session)
                labels = cluster_query.get_problem_labels(start_date, end_date)

                self.root.after(0, lambda: self._on_problem_labels_loaded(labels))
            except Exception as e:
                self.root.after(0, lambda: self._on_problem_labels_loaded([]))
                self.logger.error("[聚类工单] 加载问题标签失败: %s", str(e))

        threading.Thread(target=do_load, daemon=True).start()

    def _on_problem_labels_loaded(self, labels):
        """问题标签加载完成"""
        self.cluster_load_labels_btn.config(state=tk.NORMAL)

        if labels:
            self.logger.info("[聚类工单] 加载到 %d 个问题标签", len(labels))
            self.cluster_problem_labels = labels
            # 可以在界面上显示已加载
            self.cluster_problem_type_entry.config(foreground='#333333')
        else:
            self.logger.warning("[聚类工单] 未获取到问题标签")
            self.cluster_problem_labels = []

    def _on_cluster_query(self):
        """执行聚类工单查询"""
        if not self.session:
            messagebox.showwarning("提示", "请先登录后再操作")
            return

        if self.cluster_is_querying:
            return

        start_date = self.cluster_start_var.get()
        end_date = self.cluster_end_var.get()
        city_name = self.cluster_city_var.get()
        city_code = self.CITY_CODE_MAP.get(city_name, '')

        if not start_date or not end_date:
            messagebox.showwarning("提示", "请选择日期范围")
            return

        # 收集查询参数
        params = {
            'timeType': self.cluster_time_type_var.get(),
            'start_date': start_date,
            'end_date': end_date,
            'city': city_code,
            'area_grid': self.cluster_grid_var.get(),
            'problem_status': self.cluster_status_var.get(),
            'rows': 100,
            'page': 1
        }

        # 添加详细问题类型
        problem_type = self.cluster_problem_type_var.get().strip()
        if problem_type:
            params['detailed_type'] = [problem_type]

        self.cluster_is_querying = True
        self.cluster_query_btn.config(state=tk.DISABLED, text="查询中...")
        self.cluster_stop_btn.config(state=tk.NORMAL)

        # 清空之前的结果
        for item in self.cluster_tree.get_children():
            self.cluster_tree.delete(item)

        self.logger.info("[聚类工单] 开始查询: %s - %s", start_date, end_date)

        def do_query():
            try:
                from core.query import ClusterOrderQuery
                cluster_query = ClusterOrderQuery(self.session)

                def progress_callback(current, total, message):
                    self.root.after(0, lambda m=message: self._update_cluster_progress(m, current, total))

                result = cluster_query.query_orders(params, progress_callback)
                self.root.after(0, lambda: self._on_cluster_query_done(result))

            except Exception as e:
                self.logger.error("[聚类工单] 查询异常: %s", str(e))
                self.root.after(0, lambda: self._on_cluster_query_done({'rows': [], 'total': 0}))

        threading.Thread(target=do_query, daemon=True).start()

    def _update_cluster_progress(self, message, current, total):
        """更新聚类工单查询进度"""
        self.progress_lbl_detail.config(text=message)
        if total > 0:
            pct = int(current / total * 100) if total > 0 else 0
            self.progress_lbl_pct.config(text=f"进度: {pct}%")

    def _on_cluster_query_done(self, result):
        """聚类工单查询完成"""
        self.cluster_is_querying = False
        self.cluster_query_btn.config(state=tk.NORMAL, text="🔍 查询")
        self.cluster_stop_btn.config(state=tk.DISABLED)

        rows = result.get('rows', [])
        total = result.get('total', 0)
        total_pages = result.get('total_pages', 1)

        self.cluster_all_data = rows
        self.cluster_total_count = total
        self.cluster_total_pages = total_pages
        self.cluster_current_page = 1

        self.logger.info("[聚类工单] 查询完成: 共 %d 条数据", total)

        # 更新分页信息
        self.cluster_page_label.config(text=f"第 1 / {max(1, total_pages)} 页，共 {total} 条")

        # 显示数据
        self._display_cluster_page()

        # 启用导出按钮
        if rows:
            self.cluster_export_btn.config(state=tk.NORMAL)

    def _display_cluster_page(self):
        """显示当前页的数据"""
        # 清空表格
        for item in self.cluster_tree.get_children():
            self.cluster_tree.delete(item)

        # 显示当前页数据（每页最多100条）
        page_size = 100
        start_idx = (self.cluster_current_page - 1) * page_size
        end_idx = start_idx + page_size
        page_data = self.cluster_all_data[start_idx:end_idx]

        for i, row in enumerate(page_data):
            seq = start_idx + i + 1
            values = [
                str(seq),
                row.get('聚类工单序号', ''),
                row.get('问题小区', ''),
                row.get('问题小区名', ''),
                row.get('地市', ''),
                row.get('责任网格', ''),
                row.get('派发时间', ''),
                row.get('问题点类型', ''),
                row.get('方案类型', ''),
                row.get('方案描述', ''),
                row.get('方案确认人', ''),
                row.get('评估状态', ''),
                row.get('当前状态', ''),
                row.get('问题原因', ''),
                row.get('数据来源', '')
            ]
            self.cluster_tree.insert('', tk.END, values=values)

    def _on_cluster_prev_page(self):
        """上一页"""
        if self.cluster_current_page > 1:
            self.cluster_current_page -= 1
            self._display_cluster_page()
            self.cluster_page_label.config(
                text=f"第 {self.cluster_current_page} / {self.cluster_total_pages} 页，共 {self.cluster_total_count} 条"
            )

    def _on_cluster_next_page(self):
        """下一页"""
        if self.cluster_current_page < self.cluster_total_pages:
            self.cluster_current_page += 1
            self._display_cluster_page()
            self.cluster_page_label.config(
                text=f"第 {self.cluster_current_page} / {self.cluster_total_pages} 页，共 {self.cluster_total_count} 条"
            )

    def _on_cluster_export(self):
        """导出聚类工单数据"""
        if not self.cluster_all_data:
            messagebox.showwarning("提示", "没有数据可导出")
            return

        from tkinter import filedialog
        import pandas as pd

        file_path = filedialog.asksaveasfilename(
            title="保存聚类工单数据",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("CSV文件", "*.csv")],
            initialfile=f"聚类工单_{self.cluster_start_var.get()}_{self.cluster_end_var.get()}.xlsx"
        )

        if not file_path:
            return

        self.logger.info("[聚类工单] 正在导出数据到: %s", file_path)

        try:
            # 清理HTML内容
            import re
            def clean_html(text):
                if not isinstance(text, str):
                    return text
                # 移除HTML链接
                text = re.sub(r'<a[^>]*>([^<]*)</a>', r'\1', text)
                text = re.sub(r'<[^>]+>', '', text)
                return text.strip()

            # 构建DataFrame
            df_data = []
            for i, row in enumerate(self.cluster_all_data):
                df_data.append({
                    '序号': i + 1,
                    '聚类工单序号': clean_html(row.get('聚类工单序号', '')),
                    '问题小区': clean_html(row.get('问题小区', '')),
                    '问题小区名': clean_html(row.get('问题小区名', '')),
                    '地市': clean_html(row.get('地市', '')),
                    '责任网格': clean_html(row.get('责任网格', '')),
                    '派发时间': clean_html(row.get('派发时间', '')),
                    '问题类型': clean_html(row.get('问题点类型', '')),
                    '方案类型': clean_html(row.get('方案类型', '')),
                    '方案描述': clean_html(row.get('方案描述', '')),
                    '方案确认人': clean_html(row.get('方案确认人', '')),
                    '评估状态': clean_html(row.get('评估状态', '')),
                    '当前状态': clean_html(row.get('当前状态', '')),
                    '问题原因': clean_html(row.get('问题原因', '')),
                    '数据来源': clean_html(row.get('数据来源', ''))
                })

            df = pd.DataFrame(df_data)

            if file_path.endswith('.xlsx'):
                df.to_excel(file_path, index=False, engine='openpyxl')
            else:
                df.to_csv(file_path, index=False, encoding='utf-8-sig')

            self.logger.info("[聚类工单] 导出成功: %s (%d 条)", file_path, len(df))
            messagebox.showinfo("成功", f"导出成功！\n共 {len(df)} 条数据\n保存至: {file_path}")

        except Exception as e:
            self.logger.error("[聚类工单] 导出失败: %s", str(e))
            messagebox.showerror("错误", f"导出失败: {str(e)}")

    def _on_cluster_stop(self):
        """停止聚类工单查询"""
        self.cluster_is_querying = False
        self.cluster_query_btn.config(state=tk.NORMAL, text="🔍 查询")
        self.cluster_stop_btn.config(state=tk.DISABLED)
        self.logger.info("[聚类工单] 查询已停止")

    def _update_license_display(self):
        """更新授权时间显示"""
        if self.expiry_time:
            try:
                # 解析过期时间
                display_time = self.expiry_time.strftime("%Y-%m-%d")
                
                # 计算剩余天数
                current_dt = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
                days_left = (self.expiry_time - current_dt).days
                
                if days_left < 0:
                    self.license_label.config(text="授权已过期", fg='#fce8e6')
                elif days_left <= 7:
                    self.license_label.config(text=f"授权到期: {display_time} (剩{days_left}天)", fg='#fce8e6')
                elif days_left <= 30:
                    self.license_label.config(text=f"授权到期: {display_time} (剩{days_left}天)", fg='#fff3e0')
                else:
                    self.license_label.config(text=f"授权到期: {display_time}", fg='#e8f5e9')
            except Exception:
                self.license_label.config(text="", fg='#e8f5e9')

    def _bind_events(self):
        """绑定事件和快捷键"""
        # 绑定快捷键
        self.root.bind('<Control-l>', lambda e: self._on_login())
        self.root.bind('<Control-L>', lambda e: self._on_login())
        self.root.bind('<Control-s>', lambda e: self._on_start_export())
        self.root.bind('<Control-S>', lambda e: self._on_start_export())
        self.root.bind('<Escape>', lambda e: self._on_cancel_query() if self.is_querying else None)

        # F1 帮助
        self.root.bind('<F1>', lambda e: self._show_help())

        # 窗口关闭事件
        self.root.protocol("WM_DELETE_WINDOW", self._on_closing)

    def _show_help(self):
        """显示帮助信息"""
        help_text = """
NQI工具 - 快捷键帮助
═══════════════════════════════

Ctrl+L    - 执行登录
Ctrl+S    - 开始导出
Escape    - 取消当前查询
F1        - 显示此帮助

状态说明:
  ● 绿色  - 登录成功
  ● 黄色  - 登录中
  ● 红色  - 登录失败

提示: 点击日期框旁边的日历图标可快速选择日期
        """
        from tkinter import messagebox
        messagebox.showinfo("快捷键帮助", help_text.strip())

    def _on_start_export(self):
        """开始导出（快捷键触发）"""
        # 触发导出按钮
        self._on_export()

    def _on_cancel_query(self):
        """取消查询（快捷键触发）"""
        if self.is_querying and hasattr(self, 'jxcx') and self.jxcx:
            self.jxcx.cancel_query()
            self.log("已发送取消请求...", "WARNING")

    def _on_time_rollback(self):
        """检测到时间回拨时的处理 - 后台静默执行"""
        # 写入过期的license
        invalidate_license()
        # 强制退出程序
        self.root.after(0, self._force_exit)

    def _force_exit(self):
        """强制退出程序"""
        import os
        os._exit(1)

    def _update_login_failed_ui(self):
        """批量更新登录失败UI"""
        self.status_text.config(text="登录失败")
        self.status_dot.config(fg='#ef4444')
        self.login_status_icon.config(text="○", fg='#ef4444')
        self.login_status_lbl.config(text="登录失败", fg='#ef4444')

    def _update_login_error_ui(self, message):
        """批量更新登录异常UI"""
        self.status_text.config(text="登录异常")
        self.status_dot.config(fg='#ef4444')
        self.login_status_icon.config(text="○", fg='#ef4444')
        self.login_status_lbl.config(text="登录异常", fg='#ef4444')
        self.log(f"登录异常: {message}", "ERROR")

    def _on_category_changed(self, category):
        """数据分类选择事件"""
        self.log(f"选择了数据分类: {category}", "INFO")

    def _select_all_categories(self):
        """全选所有数据分类"""
        for var in self.category_vars.values():
            var.set(1)
        self.log("已全选所有数据分类", "INFO")

    def _deselect_all_categories(self):
        """取消全选所有数据分类"""
        for var in self.category_vars.values():
            var.set(0)
        self.log("已取消全选", "INFO")

    def _on_search_focus_in(self, event):
        """搜索框获取焦点"""
        if self.table_search_entry.get() == "搜索报表...":
            self.table_search_entry.delete(0, tk.END)
            self.table_search_entry.config(foreground='black')

    def _on_search_focus_out(self, event):
        """搜索框失去焦点"""
        if not self.table_search_entry.get():
            self.table_search_entry.insert(0, "搜索报表...")
            self.table_search_entry.config(foreground='gray')
            # 恢复所有表格
            self.table_dropdown = MultiSelectDropdown(
                None,
                self._all_tables,
                width=22,
                select_all=False
            )

    def _on_table_search(self, event):
        """搜索过滤表格"""
        search_text = self.table_search_var.get().lower()
        if search_text == "搜索报表...":
            return

        if not search_text:
            # 恢复所有表格
            self._filtered_tables = self._all_tables.copy()
        else:
            # 过滤表格
            self._filtered_tables = [t for t in self._all_tables if search_text in t.lower()]

        # 重建下拉框
        # 注意：这里需要重建下拉框内容，但由于 MultiSelectDropdown 的限制
        # 我们暂时在日志中提示搜索结果
        if search_text:
            self.log(f"搜索 '{search_text}': 找到 {len(self._filtered_tables)} 个匹配报表", "INFO")

    def _on_multi_day_toggle(self):
        """按日查询切换事件"""
        if self.multi_day_var.get():
            self.multi_day_per_sheet_cb.config(state=tk.NORMAL)
        else:
            self.multi_day_per_sheet_var.set(False)
            self.multi_day_per_sheet_cb.config(state=tk.DISABLED)

    def _on_multi_day_per_sheet_toggle(self):
        """按日分Sheet切换事件"""
        pass

    def _on_custom_fields_toggle(self):
        """自定义字段切换事件"""
        if self.custom_fields_var.get():
            self.select_fields_btn.config(state=tk.NORMAL)
        else:
            self.select_fields_btn.config(state=tk.DISABLED)

    def _on_field_mode_changed(self):
        """字段获取方式切换事件"""
        mode = self.field_mode_var.get()
        mode_text = "硬编码" if mode == 'hardcode' else "动态获取"
        self.log(f"切换字段获取方式: {mode_text}", "INFO")

    def _show_field_selector(self):
        """显示字段选择窗口"""
        # 检查登录状态
        if not self.jxcx:
            messagebox.showwarning("警告", "请先登录后再选择字段")
            return

        selected_tables = self.table_dropdown.get_selected()
        if not selected_tables:
            messagebox.showwarning("警告", "请先选择数据表")
            return

        # 创建字段选择窗口
        field_window = tk.Toplevel(self.root)
        field_window.title("选择导出字段")
        field_window.geometry("600x400")
        field_window.resizable(True, True)

        # 设置窗口在主窗口中间
        self.root.update_idletasks()
        x = (self.root.winfo_width() - 600) // 2 + self.root.winfo_x()
        y = (self.root.winfo_height() - 400) // 2 + self.root.winfo_y()
        field_window.geometry(f"600x400+{x}+{y}")

        # 创建滚动区域
        canvas = tk.Canvas(field_window)
        scrollbar = ttk.Scrollbar(field_window, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # 布局
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # 字段选择区域
        field_vars = {}
        for table_name in selected_tables:
            # 动态获取字段配置
            if table_name not in self.field_configs and self.jxcx:
                table_config = TableConfig.get_table_config(table_name)
                if table_config:
                    try:
                        self.log(f"正在获取 {table_name} 的字段配置...", "INFO")
                        configs = self.jxcx.get_field_config(
                            table_config['table_key'],
                            table_config['fieldtype'],
                            table_config['api_type']
                        )
                        if configs:
                            self.field_configs[table_name] = configs
                            self.log(f"获取到 {table_name} 的 {len(configs)} 个字段", "SUCCESS")
                        else:
                            self.log(f"获取 {table_name} 的字段配置失败，可能该报表不支持自定义字段", "WARNING")
                    except Exception as e:
                        self.log(f"获取字段配置异常: {e}", "ERROR")

            # 显示字段选择
            if table_name in self.field_configs:
                configs = self.field_configs[table_name]
                
                # 表格标题
                table_frame = tk.Frame(scrollable_frame, bg='white', bd=1, relief='solid')
                table_frame.pack(fill=tk.X, pady=5, padx=10)
                
                table_title = tk.Label(table_frame, text=table_name, 
                                     font=('Microsoft YaHei UI', 10, 'bold'),
                                     bg='white', fg='#165DFF')
                table_title.pack(anchor='w', padx=5, pady=3)

                # 字段选择
                fields_frame = tk.Frame(scrollable_frame, bg='white')
                fields_frame.pack(fill=tk.X, pady=(0, 10), padx=10)

                # 按行排列，每行4个复选框
                row_frame = None
                for i, config in enumerate(configs):
                    if i % 4 == 0:
                        row_frame = tk.Frame(fields_frame, bg='white')
                        row_frame.pack(fill=tk.X, pady=2)

                    field_name = config.get('columnname_cn', config.get('columnname', ''))
                    field_key = config.get('columnname', '')
                    
                    var = tk.BooleanVar(value=True)  # 默认全选
                    field_vars[(table_name, field_key)] = var

                    cb = tk.Checkbutton(row_frame, text=field_name,
                                       variable=var,
                                       font=('Microsoft YaHei UI', 9),
                                       bg='white', fg='#202124',
                                       selectcolor='#165DFF',
                                       activebackground='white',
                                       activeforeground='#165DFF',
                                       cursor='arrow')
                    cb.pack(side=tk.LEFT, padx=10, pady=1, fill=tk.X, expand=True)

        # 按钮区域
        btn_frame = tk.Frame(field_window, bg='white')
        btn_frame.pack(fill=tk.X, pady=10, padx=10)

        def on_ok():
            # 保存选中的字段
            self.selected_fields = {}
            for (table_name, field_key), var in field_vars.items():
                if var.get():
                    if table_name not in self.selected_fields:
                        self.selected_fields[table_name] = []
                    self.selected_fields[table_name].append(field_key)
            
            # 显示选中的字段数量
            total_fields = sum(len(fields) for fields in self.selected_fields.values())
            self.log(f"已选择 {total_fields} 个字段", "INFO")
            field_window.destroy()

        def on_cancel():
            field_window.destroy()

        ttk.Button(btn_frame, text="确定", command=on_ok, width=10).pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="取消", command=on_cancel, width=10).pack(side=tk.RIGHT, padx=5)

    def set_quick_date(self, days):
        """设置快捷日期"""
        end_date = datetime.now() - timedelta(days=1)
        start_date = end_date - timedelta(days=days-1)

        self.start_year_var.set(start_date.year)
        self.start_month_var.set(start_date.month)
        self.start_day_var.set(start_date.day)

        self.end_year_var.set(end_date.year)
        self.end_month_var.set(end_date.month)
        self.end_day_var.set(end_date.day)

        self.log(f"设置快捷日期: 近{days}天", "INFO")

    def _show_calendar(self, date_type):
        """显示日历选择对话框

        Args:
            date_type: 'start' 或 'end'，表示开始或结束日期
        """
        # 获取当前选中的日期
        if date_type == 'start':
            year = self.start_year_var.get()
            month = self.start_month_var.get()
            day = self.start_day_var.get()
            initial_date = f"{year:04d}-{month:02d}-{day:02d}"
        else:
            year = self.end_year_var.get()
            month = self.end_month_var.get()
            day = self.end_day_var.get()
            initial_date = f"{year:04d}-{month:02d}-{day:02d}"

        # 显示日历对话框
        dialog = CalendarDialog(self.root, initial_date=initial_date)
        selected_date = dialog.show()

        if selected_date:
            # 解析日期
            parts = selected_date.split('-')
            year, month, day = int(parts[0]), int(parts[1]), int(parts[2])

            # 更新变量
            if date_type == 'start':
                self.start_year_var.set(year)
                self.start_month_var.set(month)
                self.start_day_var.set(day)
                self.log(f"选择开始日期: {selected_date}", "INFO")
            else:
                self.end_year_var.set(year)
                self.end_month_var.set(month)
                self.end_day_var.set(day)
                self.log(f"选择结束日期: {selected_date}", "INFO")

    def update_progress(self, current, total, detail=""):
        """更新进度条显示

        Args:
            current: 当前进度（0-100）
            total: 总数（用于计算百分比）
            detail: 详细描述文字
        """
        if total > 0:
            pct = int(current / total * 100)
        else:
            pct = 0

        # 计算预估剩余时间
        eta_text = self._calculate_eta(current, total, pct)
        self.root.after(0, lambda: self._update_progress_ui(pct, detail, eta_text))

    def _calculate_eta(self, current, total, pct):
        """计算预估剩余时间

        Returns:
            str: 预估剩余时间文本
        """
        from datetime import datetime

        if pct <= 0:
            self._progress_start_time = datetime.now()
            self._progress_last_update = datetime.now()
            self._progress_last_value = 0
            self._total_expected = total
            return ""

        now = datetime.now()

        if self._progress_start_time is None:
            self._progress_start_time = now
            self._progress_last_update = now
            self._progress_last_value = 0
            return ""

        # 计算进度变化率
        elapsed = (now - self._progress_start_time).total_seconds()
        progress_delta = pct - self._progress_last_value

        if progress_delta > 0 and elapsed > 5:
            # 估算剩余时间
            remaining_pct = 100 - pct
            eta_seconds = (remaining_pct / progress_delta) * elapsed
            self._progress_last_update = now
            self._progress_last_value = pct

            # 格式化时间
            if eta_seconds < 60:
                eta_text = f"剩余约 {int(eta_seconds)} 秒"
            elif eta_seconds < 3600:
                minutes = int(eta_seconds / 60)
                seconds = int(eta_seconds % 60)
                eta_text = f"剩余约 {minutes}分{seconds}秒"
            else:
                hours = int(eta_seconds / 3600)
                minutes = int((eta_seconds % 3600) / 60)
                eta_text = f"剩余约 {hours}小时{minutes}分"

            return eta_text

        self._progress_last_update = now
        self._progress_last_value = pct
        return ""

    def _update_progress_ui(self, pct, detail="", eta_text=""):
        """在主线程中更新进度条UI"""
        self.progress_lbl_pct.config(text=f"进度: {pct}%")
        if detail:
            self.progress_lbl_detail.config(text=detail)
        if eta_text:
            self.progress_lbl_eta.config(text=eta_text)

        self.progress_canvas.update_idletasks()
        canvas_width = self.progress_canvas.winfo_width()
        if canvas_width <= 1:
            canvas_width = 1000

        bar_width = int(canvas_width * pct / 100)
        self.progress_canvas.coords(self.progress_bar, 0, 0, bar_width, 8)

    def reset_progress(self):
        """重置进度条"""
        self.root.after(0, lambda: self._reset_progress_ui())

    def _reset_progress_ui(self):
        """在主线程中重置进度条UI"""
        self.progress_lbl_pct.config(text="进度: 0%")
        self.progress_lbl_detail.config(text="就绪")
        self.progress_lbl_eta.config(text="")
        self.progress_canvas.coords(self.progress_bar, 0, 0, 0, 8)
        # 重置预估时间相关变量
        self._progress_start_time = None
        self._progress_last_update = None
        self._progress_last_value = 0
        self._total_expected = 0

    def open_output_dir(self):
        """打开输出目录"""
        import webbrowser
        output_dir = os.path.join(os.getcwd(), 'data_output')
        os.makedirs(output_dir, exist_ok=True)
        webbrowser.open(output_dir)

    def _on_login(self):
        """登录按钮点击事件"""
        self.log("开始登录...", "INFO")
        self.status_text.config(text="登录中...")
        self.status_dot.config(fg='#fbbf24')  # 黄色
        thread = threading.Thread(target=self._login_worker)
        thread.daemon = True
        thread.start()

    def _login_worker(self):
        """登录工作线程"""
        try:
            username = self.username_entry.get().strip()
            password = self.password_entry.get().strip()
            login_mgr = LoginManager(username=username, password=password, parent=self.root)
            if login_mgr.login():
                self.session = login_mgr.sess
                self.jxcx = JXCXQuery(self.session)
                self.root.after(0, self._on_login_success)
            else:
                self.root.after(0, self._update_login_failed_ui)
                self.root.after(0, lambda: self.log("登录失败", "ERROR"))
        except Exception as e:
            self.root.after(0, lambda msg=str(e): self._update_login_error_ui(msg))

    def _on_login_success(self):
        """登录成功回调"""
        self.log("登录成功！", "SUCCESS")
        self.extract_btn.config(state=tk.NORMAL)
        self.login_btn.config(state=tk.DISABLED, text="已登录")
        self.status_text.config(text="已登录")
        self.status_dot.config(fg='#22c55e')  # 绿色
        self.login_status_icon.config(text="●", fg='#22c55e')
        self.login_status_lbl.config(text="已登录", fg='#22c55e')

    def _on_query(self):
        """查询按钮点击事件"""
        if self.is_querying:
            self.log("正在查询中，请稍候...", "WARNING")
            return

        selected_tables = self.table_dropdown.get_selected()
        if not selected_tables:
            messagebox.showwarning("警告", "请选择要查询的数据表")
            return

        # 获取日期范围
        start_date = f"{self.start_year_var.get()}-{self.start_month_var.get():02d}-{self.start_day_var.get():02d}"
        end_date = f"{self.end_year_var.get()}-{self.end_month_var.get():02d}-{self.end_day_var.get():02d}"
        
        # 获取选中的地市
        selected_cities = self.city_dropdown.get_selected()
        city = ",".join(selected_cities) if selected_cities else ""

        self.is_querying = True
        self.extract_btn.config(state=tk.DISABLED, text="查询中...")
        self.stop_btn.config(state=tk.NORMAL)
        self.status_text.config(text="查询中...")
        self.status_dot.config(fg='#fbbf24')  # 黄色

        # 重置进度条
        self.reset_progress()
        self.update_progress(0, len(selected_tables), "开始查询...")

        self.log(f"开始查询: {', '.join(selected_tables)}", "INFO")
        self.log(f"日期范围: {start_date} 至 {end_date}", "INFO")
        if city:
            self.log(f"地市: {city}", "INFO")

        self.query_thread = threading.Thread(target=self._query_worker,
                                            args=(selected_tables, start_date, end_date, city))
        self.query_thread.daemon = True
        self.query_thread.start()

    def _on_stop(self):
        """停止查询"""
        self.log("正在停止查询...", "WARNING")
        # 调用查询模块的取消方法
        if hasattr(self, 'jxcx') and self.jxcx:
            self.jxcx.cancel_query()
        # 标记查询状态为已取消
        self.is_querying = False
        self.log("已发送取消请求，请等待当前批次完成...", "WARNING")

    def _query_worker(self, table_names, start_date, end_date, city):
        """查询工作线程"""
        try:
            from datetime import datetime, timedelta

            total_tables = len(table_names)
            multi_day = self.multi_day_var.get()
            multi_day_per_sheet = self.multi_day_per_sheet_var.get()

            for idx, table_name in enumerate(table_names):
                self.log(f"正在查询: {table_name}", "INFO")
                self.update_progress(idx, total_tables, f"正在查询: {table_name}")
                table_config = TableConfig.get_table_config(table_name)
                if not table_config:
                    self.log(f"未找到表配置: {table_name}", "ERROR")
                    continue

                self.jxcx.enter_jxcx()

                # 4G语音小区：需要分别查询VoLTE和EPSFB表后合并（跳过payload_func处理）
                is_4g_voice = table_config.get('is_4g_voice', False)
                if is_4g_voice:
                    self.log(f"4G语音小区报表：VoLTE + EPSFB 联合查询", "INFO")
                    self._query_4g_voice_table(
                        table_config, start_date, end_date, city,
                        multi_day, multi_day_per_sheet
                    )
                    self.log(f"查询完成: {table_name}", "SUCCESS")
                    continue

                # 检查是否有硬编码的payload函数
                payload_func = table_config.get('payload_func')
                if payload_func:
                    # 检查是否是工参报表（使用__gongcan__标记）
                    # 先获取payload模板看看是否为工参报表
                    payload_template = payload_func()
                    if payload_template and payload_template.get('__gongcan__'):
                        # 工参报表需要特殊处理（不需要时间条件）
                        self.log(f"工参报表: 使用table类型API", "INFO")
                        conditions = table_config.get('default_conditions', []).copy()
                        if city:
                            conditions.append({'field': 'city', 'operator': 'in', 'value': city})

                        gongcan_payload = self.jxcx.build_payload_from_config(
                            payload_template.get('table_key'),
                            payload_template.get('fieldtype'),
                            conditions,
                            payload_template.get('api_type', 'table'),
                            dimension_override={
                                'geographicdimension': payload_template.get('geographicdimension', ''),
                                'timedimension': payload_template.get('timedimension', ''),
                                'enodebField': payload_template.get('enodebField', ''),
                                'cgiField': payload_template.get('cgiField', ''),
                                'timeField': payload_template.get('timeField', ''),
                                'cellField': payload_template.get('cellField', ''),
                                'cityField': payload_template.get('cityField', '')
                            }
                        )
                        if gongcan_payload:
                            df = self.jxcx.get_table(gongcan_payload, report_name=table_name)
                            if not df.empty:
                                filename = f"{table_name}.xlsx"
                                filepath = export_with_format(df, filename, table_name)
                                if filepath:
                                    self.log(f"数据已导出到: {os.path.basename(filepath)}", "SUCCESS")
                                else:
                                    self.log(f"导出失败: {table_name}", "ERROR")
                            else:
                                self.log(f"查询结果为空: {table_name}", "WARNING")
                        self.log(f"查询完成: {table_name}", "SUCCESS")
                        continue

                    # 按日查询模式
                    if multi_day:
                        # 按日查询模式：每天分别查询
                        current_date = datetime.strptime(start_date, '%Y-%m-%d')
                        end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
                        dates = []
                        while current_date <= end_datetime:
                            dates.append(current_date.strftime('%Y-%m-%d'))
                            current_date += timedelta(days=1)

                        total_days = len(dates)
                        self.log(f"按日查询模式: 共 {total_days} 天", "INFO")

                        all_dfs = []  # 用于收集所有日数据

                        for day_idx, query_date in enumerate(dates):
                            self.log(f"查询 {table_name} [{day_idx + 1}/{total_days}]: {query_date}", "INFO")
                            self.update_progress(idx + day_idx / total_days, total_tables,
                                               f"查询 {table_name}: {query_date}")

                            # 调用payload_func，传入单日日期
                            payload = payload_func(query_date, query_date, city)

                            if payload:
                                # 调试日志：输出where条件中的日期
                                if 'where' in payload:
                                    for cond in payload['where']:
                                        if 'starttime' in cond.get('feild', '') and 'time' in cond.get('feild', '').lower():
                                            self.log(f"  [调试] 日期条件: {cond.get('feild')} {cond.get('symbol')} {cond.get('val')}", "INFO")
                                df = self.jxcx.get_table(payload, report_name=table_name)
                                if not df.empty:
                                    # 检查是否需要添加计算列（全程完好率报表）
                                    calc_columns = table_config.get('calc_columns', [])
                                    add_calc = bool(calc_columns)

                                    if multi_day_per_sheet:
                                        # 按日分Sheet模式：每天一个Sheet
                                        if add_calc:
                                            try:
                                                if '4G全程完好率' in table_name:
                                                    df = self._add_4g_wanchenglv_calc_columns(df)
                                                elif '5G全程完好率' in table_name:
                                                    df = self._add_5g_wanchenglv_calc_columns(df)
                                                self.log(f"  [计算列] {query_date} 计算列添加完成", "SUCCESS")
                                            except Exception as e:
                                                self.log(f"  [计算列] {query_date} 计算列添加异常: {e}", "WARNING")
                                        all_dfs.append((query_date.replace('-', ''), df))
                                    else:
                                        # 按日分文件模式：每天一个文件
                                        if add_calc:
                                            try:
                                                if '4G全程完好率' in table_name:
                                                    df = self._add_4g_wanchenglv_calc_columns(df)
                                                elif '5G全程完好率' in table_name:
                                                    df = self._add_5g_wanchenglv_calc_columns(df)
                                                self.log(f"  [计算列] {query_date} 计算列添加完成", "SUCCESS")
                                            except Exception as e:
                                                self.log(f"  [计算列] {query_date} 计算列添加异常: {e}", "WARNING")
                                        day_filename = f"{table_name}_{query_date}.xlsx"
                                        day_filepath = export_with_format(df, day_filename, table_name)
                                        if day_filepath:
                                            self.log(f"  {query_date}: {len(df)} 条数据 -> {os.path.basename(day_filepath)}", "SUCCESS")
                                        all_dfs.append(df)

                        # 按日分Sheet模式：合并所有日到同一文件
                        if multi_day_per_sheet and all_dfs:
                            day_filename = f"{table_name}_{start_date}_{end_date}.xlsx"
                            self._export_multi_sheet(day_filename, all_dfs, table_name)
                            self.log(f"按日分Sheet导出完成: {day_filename}", "SUCCESS")

                        # 按日分文件模式：合并所有日到同一文件
                        if not multi_day_per_sheet and all_dfs:
                            combined_df = pd.concat(all_dfs, ignore_index=True)
                            day_filename = f"{table_name}_{start_date}_{end_date}.xlsx"
                            combined_filepath = export_with_format(combined_df, day_filename, table_name)
                            if combined_filepath:
                                self.log(f"按日查询导出完成: {os.path.basename(combined_filepath)} ({len(combined_df)} 条)", "SUCCESS")

                        if not all_dfs:
                            self.log(f"查询结果为空: {table_name}", "WARNING")
                    else:
                        # 普通模式：按日期范围查询
                        self.log(f"使用硬编码payload模板: {table_name}", "INFO")
                        self.log(f"  [调试] 查询日期范围: {start_date} 至 {end_date}", "INFO")
                        payload = payload_func(start_date, end_date, city)

                        if payload:
                            # 调试日志：输出where条件中的日期
                            if 'where' in payload:
                                for cond in payload['where']:
                                    if 'starttime' in cond.get('feild', ''):
                                        self.log(f"  [调试] 日期条件: {cond.get('feild')} {cond.get('symbol')} {cond.get('val')}", "INFO")
                            df = self.jxcx.get_table(payload, report_name=table_name)
                            if not df.empty:
                                # 检查是否需要添加计算列（全程完好率报表）
                                calc_columns = table_config.get('calc_columns', [])
                                if calc_columns:
                                    self.log(f"[计算列] 开始为 {table_name} 添加计算列: {calc_columns}", "INFO")
                                    try:
                                        if '4G全程完好率' in table_name:
                                            df = self._add_4g_wanchenglv_calc_columns(df)
                                        elif '5G全程完好率' in table_name:
                                            df = self._add_5g_wanchenglv_calc_columns(df)
                                        self.log(f"[计算列] {table_name} 计算列添加完成", "SUCCESS")
                                    except Exception as e:
                                        self.log(f"[计算列] {table_name} 计算列添加异常: {e}", "ERROR")

                                filename = f"{table_name}_{start_date}_{end_date}.xlsx"
                                filepath = export_with_format(df, filename, table_name)
                                if filepath:
                                    self.log(f"数据已导出到: {os.path.basename(filepath)}", "SUCCESS")
                                else:
                                    self.log(f"导出失败: {table_name}", "ERROR")
                            else:
                                self.log(f"查询结果为空: {table_name}", "WARNING")

                    self.log(f"查询完成: {table_name}", "SUCCESS")
                    continue

                # 获取维度参数和字段配置
                dimension = table_config.get('dimension', {})
                # 根据用户选择的模式决定字段获取方式
                field_mode = getattr(self, 'field_mode_var', None)
                use_hardcode_fields = field_mode.get() == 'hardcode' if field_mode else True
                # 如果使用硬编码模式且有预定义字段，则使用；否则动态获取
                fields = table_config.get('fields', None) if use_hardcode_fields else None
                if use_hardcode_fields and fields:
                    self.log(f"使用硬编码字段配置 (共 {len(fields)} 个字段)", "INFO")
                elif use_hardcode_fields:
                    self.log("使用硬编码模式但未找到预定义字段，将动态获取", "WARNING")
                else:
                    self.log("使用动态字段获取模式", "INFO")
                is_gongcan = table_config.get('is_gongcan', False)
                is_4g_voice = table_config.get('is_4g_voice', False)

                # 4G语音小区：需要分别查询VoLTE和EPSFB表后合并
                if is_4g_voice:
                    self.log(f"4G语音小区报表：VoLTE + EPSFB 联合查询", "INFO")
                    self._query_4g_voice_table(
                        table_config, start_date, end_date, city,
                        multi_day, multi_day_per_sheet
                    )
                    self.log(f"查询完成: {table_name}", "SUCCESS")
                    continue

                # 工参报表不需要时间条件，直接按原逻辑处理
                if is_gongcan:
                    conditions = table_config.get('default_conditions', []).copy()
                    if city:
                        conditions.append({'field': 'city', 'operator': 'in', 'value': city})

                    payload = self.jxcx.build_payload_from_config(
                        table_config['table_key'],
                        table_config['fieldtype'],
                        conditions,
                        table_config['api_type'],
                        dimension_override=dimension if dimension else None,
                        fields_override=fields,
                        table_name=table_config.get('table_name')
                    )

                    if payload:
                        df = self.jxcx.get_table(payload, report_name=table_name)
                        if not df.empty:
                            self._apply_custom_fields(df, table_name)
                            filename = f"{table_name}_{start_date}_{end_date}.xlsx"
                            filepath = export_with_format(df, filename, table_name)
                            if filepath:
                                self.log(f"数据已导出到: {os.path.basename(filepath)}", "SUCCESS")
                            else:
                                self.log(f"导出失败: {table_name}", "ERROR")
                        else:
                            self.log(f"查询结果为空: {table_name}", "WARNING")
                    self.log(f"查询完成: {table_name}", "SUCCESS")
                    continue

                # 非工参报表：按日查询处理
                if multi_day:
                    # 按日查询模式：每天分别查询
                    current_date = datetime.strptime(start_date, '%Y-%m-%d')
                    end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
                    dates = []
                    while current_date <= end_datetime:
                        dates.append(current_date.strftime('%Y-%m-%d'))
                        current_date += timedelta(days=1)

                    total_days = len(dates)
                    self.log(f"按日查询模式: 共 {total_days} 天", "INFO")

                    all_dfs = []  # 用于收集所有日数据
                    per_sheet_writer = None  # 用于按日分Sheet

                    for day_idx, query_date in enumerate(dates):
                        self.log(f"查询 {table_name} [{day_idx + 1}/{total_days}]: {query_date}", "INFO")
                        self.update_progress(idx + day_idx / total_days, total_tables,
                                           f"查询 {table_name}: {query_date}")

                        conditions = table_config.get('default_conditions', []).copy()
                        conditions.append({'field': 'starttime', 'operator': '>=', 'value': query_date})
                        conditions.append({'field': 'starttime', 'operator': '<=', 'value': query_date})
                        if city:
                            conditions.append({'field': 'city', 'operator': 'in', 'value': city})

                        payload = self.jxcx.build_payload_from_config(
                            table_config['table_key'],
                            table_config['fieldtype'],
                            conditions,
                            table_config['api_type'],
                            dimension_override=dimension if dimension else None,
                            fields_override=fields,
                            table_name=table_config.get('table_name')
                        )

                        if payload:
                            df = self.jxcx.get_table(payload, report_name=table_name)
                            if not df.empty:
                                if multi_day_per_sheet:
                                    # 按日分Sheet模式：每天一个Sheet
                                    day_filename = f"{table_name}_{start_date}_{end_date}.xlsx"
                                    day_filepath = export_with_format(
                                        df, day_filename,
                                        sheet_name=query_date.replace('-', '')
                                    )
                                    if day_filepath:
                                        self.log(f"  {query_date}: {len(df)} 条数据 -> Sheet", "SUCCESS")
                                    all_dfs.append((query_date, df))
                                else:
                                    # 按日分文件模式：每天一个文件
                                    day_filename = f"{table_name}_{query_date}.xlsx"
                                    day_filepath = export_with_format(df, day_filename, table_name)
                                    if day_filepath:
                                        self.log(f"  {query_date}: {len(df)} 条数据 -> {os.path.basename(day_filepath)}", "SUCCESS")
                                    all_dfs.append(df)

                    # 按日分Sheet模式：合并所有日到同一文件
                    if multi_day_per_sheet and all_dfs:
                        day_filename = f"{table_name}_{start_date}_{end_date}.xlsx"
                        self._export_multi_sheet(day_filename, all_dfs, table_name)
                        self.log(f"按日分Sheet导出完成: {day_filename}", "SUCCESS")

                    # 按日分文件模式：合并所有日到同一文件
                    if not multi_day_per_sheet and all_dfs:
                        combined_df = pd.concat(all_dfs, ignore_index=True)
                        day_filename = f"{table_name}_{start_date}_{end_date}.xlsx"
                        combined_filepath = export_with_format(combined_df, day_filename, table_name)
                        if combined_filepath:
                            self.log(f"按日查询导出完成: {os.path.basename(combined_filepath)} ({len(combined_df)} 条)", "SUCCESS")

                    if not all_dfs:
                        self.log(f"查询结果为空: {table_name}", "WARNING")

                else:
                    # 普通模式：按日期范围查询（原有逻辑）
                    conditions = table_config.get('default_conditions', []).copy()
                    conditions.append({'field': 'starttime', 'operator': '>=', 'value': start_date})
                    conditions.append({'field': 'starttime', 'operator': '<=', 'value': end_date})
                    if city:
                        conditions.append({'field': 'city', 'operator': 'in', 'value': city})

                    payload = self.jxcx.build_payload_from_config(
                        table_config['table_key'],
                        table_config['fieldtype'],
                        conditions,
                        table_config['api_type'],
                        dimension_override=dimension if dimension else None,
                        fields_override=fields,
                        table_name=table_config.get('table_name')
                    )

                    if payload:
                        df = self.jxcx.get_table(payload, report_name=table_name)
                        if not df.empty:
                            self._apply_custom_fields(df, table_name)
                            filename = f"{table_name}_{start_date}_{end_date}.xlsx"
                            filepath = export_with_format(df, filename, table_name)
                            if filepath:
                                self.log(f"数据已导出到: {os.path.basename(filepath)}", "SUCCESS")
                            else:
                                self.log(f"导出失败: {table_name}", "ERROR")
                        else:
                            self.log(f"查询结果为空: {table_name}", "WARNING")

                self.log(f"查询完成: {table_name}", "SUCCESS")

            self.root.after(0, self._on_query_complete)
        except Exception as e:
            self.root.after(0, lambda: self.log(f"查询异常: {e}", "ERROR"))
            import traceback
            traceback.print_exc()
            self.root.after(0, self._on_query_failed)

    def _apply_custom_fields(self, df, table_name):
        """应用自定义字段选择到 DataFrame"""
        if self.custom_fields_var.get() and table_name in self.selected_fields:
            selected_field_keys = self.selected_fields[table_name]
            available_fields = [col for col in df.columns if col in selected_field_keys]
            if not available_fields and table_name in self.field_configs:
                config_map = {c.get('columnname'): c.get('columnname') for c in self.field_configs[table_name]}
                for col in df.columns:
                    if col in config_map and config_map[col] in selected_field_keys:
                        available_fields.append(col)
            if available_fields:
                df = df[available_fields]
                self.log(f"应用自定义字段: {len(available_fields)} 个字段", "INFO")
            else:
                self.log(f"没有选中的字段可用，将导出全部字段", "WARNING")
        return df

    def _export_multi_sheet(self, filename, sheets_data, default_sheet):
        """导出多Sheet的Excel文件

        Args:
            filename: 文件名
            sheets_data: list of (sheet_name, DataFrame) tuples
            default_sheet: 默认工作表名称
        """
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        ensure_dirs()
        filepath = os.path.join(OUTPUT_DIR, filename)

        wb = Workbook()
        wb.remove(wb.active)  # 移除默认sheet

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='165DFF', end_color='165DFF', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for sheet_name, df in sheets_data:
            ws = wb.create_sheet(title=str(sheet_name)[:31])  # Sheet名最多31字符
            ws.append(list(df.columns))

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            for row in df.itertuples(index=False):
                ws.append(list(row))

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (AttributeError, TypeError):
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        wb.save(filepath)

    def _query_4g_voice_table(self, table_config, start_date, end_date, city,
                               multi_day, multi_day_per_sheet):
        """查询4G语音小区报表（VoLTE + EPSFB 联合查询）

        Args:
            table_config: 表配置
            start_date: 开始日期
            end_date: 结束日期
            city: 地市
            multi_day: 是否按日查询
            multi_day_per_sheet: 是否按日分Sheet
        """
        from datetime import datetime, timedelta
        import numpy as np

        volte_fields = table_config.get('volte_fields', [])
        epsfb_fields = table_config.get('epsfb_fields', [])

        if not volte_fields or not epsfb_fields:
            self.log(f"4G语音小区字段配置不完整，无法查询", "ERROR")
            return

        # VoLTE和EPSFB的维度配置
        volte_dimension = table_config.get('dimension', {})
        epsfb_dimension = {
            'geographicdimension': '小区',
            'timedimension': '天',
            'enodebField': '---',  # EPSFB表没有enodeb字段
            'cgiField': 'cgi',
            'timeField': 'starttime',
            'cellField': 'cell',
            'cityField': 'city',
        }

        # 按日查询模式
        if multi_day:
            current_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
            dates = []
            while current_date <= end_datetime:
                dates.append(current_date.strftime('%Y-%m-%d'))
                current_date += timedelta(days=1)

            total_days = len(dates)
            self.log(f"4G语音小区 - 按日查询模式: 共 {total_days} 天", "INFO")

            all_merged_dfs = []

            for day_idx, query_date in enumerate(dates):
                self.log(f"4G语音小区 [{day_idx + 1}/{total_days}]: {query_date}", "INFO")
                self.update_progress(day_idx, total_days, f"查询4G语音小区: {query_date}")

                # 构建VoLTE和EPSFB的查询条件
                conditions = [
                    {'field': 'starttime', 'operator': '>=', 'value': query_date},
                    {'field': 'starttime', 'operator': '<=', 'value': query_date},
                ]
                if city:
                    conditions.append({'field': 'city', 'operator': 'in', 'value': city})

                # 构建payload
                payloads = self.jxcx.build_4g_voice_payload(
                    volte_fields, epsfb_fields, conditions,
                    volte_dimension, epsfb_dimension
                )
                volte_payload = payloads['volte']
                epsfb_payload = payloads['epsfb']

                # 获取数据
                voice_data = self.jxcx.get_4g_voice_table(volte_payload, epsfb_payload)
                merged_df = voice_data

                if not merged_df.empty:
                    # 添加计算列
                    try:
                        merged_df = self._add_4g_voice_calc_columns(merged_df)
                    except Exception as e:
                        self.log(f"  添加计算列异常: {e}", "WARNING")

                    if multi_day_per_sheet:
                        # 按日分Sheet模式：每天一个Sheet
                        all_merged_dfs.append((query_date.replace('-', ''), merged_df))
                    else:
                        # 按日分文件模式：每天一个文件
                        filename = f"4G语音小区_{query_date}.xlsx"
                        filepath = export_with_format(merged_df, filename, "4G语音小区")
                        if filepath:
                            self.log(f"  {query_date}: {len(merged_df)} 条数据 -> {os.path.basename(filepath)}", "SUCCESS")

            # 按日分Sheet模式：合并所有日到同一文件
            if multi_day_per_sheet and all_merged_dfs:
                filename = f"4G语音小区_{start_date}_{end_date}.xlsx"
                self._export_multi_sheet(filename, all_merged_dfs, "4G语音小区")
                self.log(f"按日分Sheet导出完成: {filename}", "SUCCESS")

        else:
            # 普通模式：按日期范围查询
            conditions = [
                {'field': 'starttime', 'operator': '>=', 'value': start_date},
                {'field': 'starttime', 'operator': '<=', 'value': end_date},
            ]
            if city:
                conditions.append({'field': 'city', 'operator': 'in', 'value': city})

            self.log(f"4G语音小区: 查询 {start_date} 至 {end_date}", "INFO")

            # 构建payload
            payloads = self.jxcx.build_4g_voice_payload(
                volte_fields, epsfb_fields, conditions,
                volte_dimension, epsfb_dimension
            )
            volte_payload = payloads['volte']
            epsfb_payload = payloads['epsfb']

            # 获取数据
            merged_df = self.jxcx.get_4g_voice_table(volte_payload, epsfb_payload)

            if not merged_df.empty:
                # 添加计算列
                try:
                    merged_df = self._add_4g_voice_calc_columns(merged_df)
                except Exception as e:
                    self.log(f"添加计算列异常: {e}", "WARNING")

                filename = f"4G语音小区_{start_date}_{end_date}.xlsx"
                filepath = export_with_format(merged_df, filename, "4G语音小区")
                if filepath:
                    self.log(f"4G语音小区数据已导出到: {os.path.basename(filepath)}", "SUCCESS")
            else:
                self.log(f"4G语音小区查询结果为空", "WARNING")

    def _add_4g_voice_calc_columns(self, df):
        """添加4G语音小区计算列（预警报表版本）

        计算规则：
        - 4G语音通话质差时长比例 = (VoLTE语音上行吞字时长+VoLTE语音上行单通时长+VoLTE语音上行断续时长+EPSFB语音上行吞字时长+EPSFB语音上行单通时长+EPSFB语音上行断续时长) / (VoLTE语音上行总时长+EPSFB语音上行总时长)
        - 4G差小区 = (VoLTE语音通话质差时长比例>2% 且 VoLTE通话次数>1000) 或者 (EPSFB语音通话质差时长比例>2% 且 EPSFB通话次数>1000)

        Args:
            df: 合并后的DataFrame

        Returns:
            DataFrame: 添加了计算列的DataFrame
        """
        import numpy as np

        # 获取VoLTE和EPSFB的字段名（预警报表使用中文名）
        volte_ul_tunzi = None
        volte_ul_dantong = None
        volte_ul_duanxu = None
        volte_ul_sum = None
        volte_call = None
        epsfb_ul_tunzi = None
        epsfb_ul_dantong = None
        epsfb_ul_duanxu = None
        epsfb_ul_sum = None
        epsfb_call = None

        for col in df.columns:
            col_lower = col.lower()
            # 预警报表的中文字段名
            if 'volte' in col_lower and '上行' in col and '吞字' in col:
                volte_ul_tunzi = col
            elif 'volte' in col_lower and '上行' in col and '单通' in col:
                volte_ul_dantong = col
            elif 'volte' in col_lower and '上行' in col and '断续' in col:
                volte_ul_duanxu = col
            elif 'volte' in col_lower and '上行' in col and '总时长' in col:
                volte_ul_sum = col
            elif 'volte' in col_lower and ('通话次数' in col or '语音通话总次数' in col):
                volte_call = col
            elif 'epsfb' in col_lower and '上行' in col and '吞字' in col:
                epsfb_ul_tunzi = col
            elif 'epsfb' in col_lower and '上行' in col and '单通' in col:
                epsfb_ul_dantong = col
            elif 'epsfb' in col_lower and '上行' in col and '断续' in col:
                epsfb_ul_duanxu = col
            elif 'epsfb' in col_lower and '上行' in col and '总时长' in col:
                epsfb_ul_sum = col
            elif 'epsfb' in col_lower and ('通话次数' in col or '语音通话总次数' in col):
                epsfb_call = col

        # 如果没找到中文列名，尝试使用英文列名
        if volte_ul_tunzi is None:
            for col in df.columns:
                if 'volte_ul_tunzi' in col.lower():
                    volte_ul_tunzi = col
                    break
        if volte_ul_dantong is None:
            for col in df.columns:
                if 'volte_ul_dantong' in col.lower():
                    volte_ul_dantong = col
                    break
        if volte_ul_duanxu is None:
            for col in df.columns:
                if 'volte_ul_duanxu' in col.lower():
                    volte_ul_duanxu = col
                    break
        if volte_ul_sum is None:
            for col in df.columns:
                if 'volte_ul_voice_sum' in col.lower():
                    volte_ul_sum = col
                    break
        if volte_call is None:
            for col in df.columns:
                if 'volte_ans_voice' in col.lower():
                    volte_call = col
                    break
        if epsfb_ul_tunzi is None:
            for col in df.columns:
                if 'epsfb_ul_tunzi' in col.lower():
                    epsfb_ul_tunzi = col
                    break
        if epsfb_ul_dantong is None:
            for col in df.columns:
                if 'epsfb_ul_dantong' in col.lower():
                    epsfb_ul_dantong = col
                    break
        if epsfb_ul_duanxu is None:
            for col in df.columns:
                if 'epsfb_ul_duanxu' in col.lower():
                    epsfb_ul_duanxu = col
                    break
        if epsfb_ul_sum is None:
            for col in df.columns:
                if 'epsfb_ul_voice_sum' in col.lower():
                    epsfb_ul_sum = col
                    break
        if epsfb_call is None:
            for col in df.columns:
                if 'epsfb_ans_voice' in col.lower():
                    epsfb_call = col
                    break

        # 记录找到的字段
        found_fields = []
        missing_fields = []
        if volte_ul_tunzi:
            found_fields.append(f"VoLTE上行吞字={volte_ul_tunzi[:20]}...")
        else:
            missing_fields.append("VoLTE上行吞字")
        if volte_ul_dantong:
            found_fields.append(f"VoLTE上行单通={volte_ul_dantong[:20]}...")
        else:
            missing_fields.append("VoLTE上行单通")
        if volte_ul_duanxu:
            found_fields.append(f"VoLTE上行断续={volte_ul_duanxu[:20]}...")
        else:
            missing_fields.append("VoLTE上行断续")
        if volte_ul_sum:
            found_fields.append(f"VoLTE上行总时长={volte_ul_sum[:20]}...")
        else:
            missing_fields.append("VoLTE上行总时长")
        if volte_call:
            found_fields.append(f"VoLTE通话次数={volte_call[:20]}...")
        else:
            missing_fields.append("VoLTE通话次数")
        if epsfb_ul_tunzi:
            found_fields.append(f"EPSFB上行吞字={epsfb_ul_tunzi[:20]}...")
        else:
            missing_fields.append("EPSFB上行吞字")
        if epsfb_ul_dantong:
            found_fields.append(f"EPSFB上行单通={epsfb_ul_dantong[:20]}...")
        else:
            missing_fields.append("EPSFB上行单通")
        if epsfb_ul_duanxu:
            found_fields.append(f"EPSFB上行断续={epsfb_ul_duanxu[:20]}...")
        else:
            missing_fields.append("EPSFB上行断续")
        if epsfb_ul_sum:
            found_fields.append(f"EPSFB上行总时长={epsfb_ul_sum[:20]}...")
        else:
            missing_fields.append("EPSFB上行总时长")
        if epsfb_call:
            found_fields.append(f"EPSFB通话次数={epsfb_call[:20]}...")
        else:
            missing_fields.append("EPSFB通话次数")

        self.log(f"[4G语音计算] 找到字段: {len(found_fields)}, 缺失: {len(missing_fields)}", "INFO")
        if missing_fields:
            self.log(f"[4G语音计算] 缺失字段: {missing_fields}", "WARNING")

        # 计算4G语音通话质差时长比例
        # = (VoLTE语音上行吞字时长+VoLTE语音上行单通时长+VoLTE语音上行断续时长+EPSFB语音上行吞字时长+EPSFB语音上行单通时长+EPSFB语音上行断续时长)
        #   / (VoLTE语音上行总时长+EPSFB语音上行总时长)
        if all(v is not None for v in [volte_ul_tunzi, volte_ul_dantong, volte_ul_duanxu,
                                         volte_ul_sum, epsfb_ul_tunzi, epsfb_ul_dantong,
                                         epsfb_ul_duanxu, epsfb_ul_sum]):
            total_bad = (df[volte_ul_tunzi].fillna(0) + df[volte_ul_dantong].fillna(0) +
                         df[volte_ul_duanxu].fillna(0) + df[epsfb_ul_tunzi].fillna(0) +
                         df[epsfb_ul_dantong].fillna(0) + df[epsfb_ul_duanxu].fillna(0))
            total_sum = df[volte_ul_sum].fillna(0) + df[epsfb_ul_sum].fillna(0)
            df['4G语音通话质差时长比例'] = np.where(total_sum > 0, (total_bad / total_sum * 100).round(4), np.nan)
            self.log(f"[4G语音计算] 4G语音通话质差时长比例计算完成", "SUCCESS")
        else:
            self.log(f"[4G语音计算] 缺少必要字段，无法计算4G语音通话质差时长比例", "WARNING")

        # 计算4G语音差小区判定
        # 条件：(VoLTE语音通话质差时长比例>2%且VoLTE通话次数>1000)
        # 或者 (EPSFB语音通话质差时长比例>2%且EPSFB通话次数>1000)
        volte_bad_rate = None
        epsfb_bad_rate = None

        if volte_ul_sum and volte_ul_tunzi and volte_ul_dantong and volte_ul_duanxu:
            volte_total_bad = df[volte_ul_tunzi].fillna(0) + df[volte_ul_dantong].fillna(0) + df[volte_ul_duanxu].fillna(0)
            volte_total = df[volte_ul_sum].fillna(0)
            volte_bad_rate = np.where(volte_total > 0, volte_total_bad / volte_total * 100, 0)

        if epsfb_ul_sum and epsfb_ul_tunzi and epsfb_ul_dantong and epsfb_ul_duanxu:
            epsfb_total_bad = df[epsfb_ul_tunzi].fillna(0) + df[epsfb_ul_dantong].fillna(0) + df[epsfb_ul_duanxu].fillna(0)
            epsfb_total = df[epsfb_ul_sum].fillna(0)
            epsfb_bad_rate = np.where(epsfb_total > 0, epsfb_total_bad / epsfb_total * 100, 0)

        if volte_bad_rate is not None and volte_call is not None:
            volte_is_bad = (volte_bad_rate > 2) & (df[volte_call].fillna(0) > 1000)
            bad_volte_count = volte_is_bad.sum()
            self.log(f"[4G语音计算] VoLTE差小区数量: {bad_volte_count}", "INFO")
        else:
            volte_is_bad = pd.Series([False] * len(df))
            self.log(f"[4G语音计算] 无法计算VoLTE差小区（缺少字段）", "WARNING")

        if epsfb_bad_rate is not None and epsfb_call is not None:
            epsfb_is_bad = (epsfb_bad_rate > 2) & (df[epsfb_call].fillna(0) > 1000)
            bad_epsfb_count = epsfb_is_bad.sum()
            self.log(f"[4G语音计算] EPSFB差小区数量: {bad_epsfb_count}", "INFO")
        else:
            epsfb_is_bad = pd.Series([False] * len(df))
            self.log(f"[4G语音计算] 无法计算EPSFB差小区（缺少字段）", "WARNING")

        df['4G语音差小区'] = np.where(volte_is_bad | epsfb_is_bad, '是', '否')
        bad_cell_count = (df['4G语音差小区'] == '是').sum()
        self.log(f"[4G语音计算] 4G语音差小区总计: {bad_cell_count}", "SUCCESS")

        return df

    def _add_4g_wanchenglv_calc_columns(self, df):
        """添加4G全程完好率计算列

        计算规则：
        - 4G无线接通率(%) = (RRC连接建立成功次数/RRC连接建立请求次数) * (E-RAB建立成功数/E-RAB建立请求数)
        - 4G切换成功率(%) = 切换成功次数/切换请求次数
        - 4G E-RAB掉线率 = (切出失败的E-RAB数 - 正常的eNB请求释放的E-RAB数 + eNB请求释放的E-RAB数)
                           / (遗留上下文个数 + E-RAB建立成功数 + 切换入E-RAB数)
        - 4G全程完好率 = 4G无线接通率(%) * 4G切换成功率(%) * (100 - 4G E-RAB掉线率)
        - 4G是否差小区 = 全程完好率 < 85% 时为"是"，否则为"否"

        Args:
            df: 4G全程完好率报表的DataFrame

        Returns:
            DataFrame: 添加了计算列的DataFrame
        """
        import numpy as np

        # 字段名映射（英文 -> 中文/通用名）
        field_map = {}
        for col in df.columns:
            col_lower = col.lower() if isinstance(col, str) else ''
            # RRC连接建立成功次数
            if col_lower in ('succconnestab', 'rrc连接建立成功次数'):
                field_map['rrc_succ'] = col
            # RRC连接建立请求次数
            elif col_lower in ('attconnestab', 'rrc连接建立请求次数'):
                field_map['rrc_att'] = col
            # E-RAB建立成功数
            elif col_lower in ('nbrsuccestab', 'e-rab建立成功数'):
                field_map['erab_succ'] = col
            # E-RAB建立请求数
            elif col_lower in ('nbrattestab', 'e-rab建立请求数'):
                field_map['erab_att'] = col
            # 切换成功次数
            elif col_lower in ('ho_succ_out', '切换成功次数'):
                field_map['ho_succ'] = col
            # 切换请求次数
            elif col_lower in ('ho_att__out', '切换请求次数'):
                field_map['ho_att'] = col
            # 切出失败的E-RAB数
            elif col_lower in ('hofail', '切出失败的e-rab数'):
                field_map['ho_fail'] = col
            # 正常的eNB请求释放的E-RAB数
            elif col_lower in ('nbrreqrelenb_normal', '正常的enb请求释放的e-rab数'):
                field_map['erab_normal_rel'] = col
            # eNB请求释放的E-RAB数
            elif col_lower in ('nbrreqrelenb', 'enb请求释放的e-rab数'):
                field_map['erab_rel'] = col
            # 遗留上下文个数
            elif col_lower in ('nbrleft', '遗留上下文个数'):
                field_map['context_left'] = col
            # 切换入E-RAB数
            elif col_lower in ('nbrhoinc', '切换入e-rab数'):
                field_map['ho_inc'] = col

        self.log(f"[4G全程完好率] 字段映射: {list(field_map.keys())}", "INFO")

        # ========== 1. 计算4G无线接通率(%) ==========
        if 'rrc_succ' in field_map and 'rrc_att' in field_map and 'erab_succ' in field_map and 'erab_att' in field_map:
            rrc_succ = df[field_map['rrc_succ']].fillna(0).astype(float)
            rrc_att = df[field_map['rrc_att']].fillna(0).astype(float)
            erab_succ = df[field_map['erab_succ']].fillna(0).astype(float)
            erab_att = df[field_map['erab_att']].fillna(0).astype(float)

            rrc_rate = np.where(rrc_att > 0, rrc_succ / rrc_att * 100, np.nan)
            erab_rate = np.where(erab_att > 0, erab_succ / erab_att * 100, np.nan)
            df['4G无线接通率(%)'] = np.where(
                (rrc_att > 0) & (erab_att > 0),
                rrc_rate * erab_rate,
                np.nan
            )
            df['4G无线接通率(%)'] = df['4G无线接通率(%)'].round(4)
            valid_count = df['4G无线接通率(%)'].notna().sum()
            self.log(f"[4G全程完好率] 4G无线接通率计算完成, 有效数据: {valid_count}", "SUCCESS")
        else:
            missing = [k for k in ['rrc_succ', 'rrc_att', 'erab_succ', 'erab_att'] if k not in field_map]
            self.log(f"[4G全程完好率] 缺少字段无法计算4G无线接通率: {missing}", "WARNING")

        # ========== 2. 计算4G切换成功率(%) ==========
        if 'ho_succ' in field_map and 'ho_att' in field_map:
            ho_succ = df[field_map['ho_succ']].fillna(0).astype(float)
            ho_att = df[field_map['ho_att']].fillna(0).astype(float)
            df['4G切换成功率(%)'] = np.where(ho_att > 0, ho_succ / ho_att * 100, np.nan)
            df['4G切换成功率(%)'] = df['4G切换成功率(%)'].round(4)
            valid_count = df['4G切换成功率(%)'].notna().sum()
            self.log(f"[4G全程完好率] 4G切换成功率计算完成, 有效数据: {valid_count}", "SUCCESS")
        else:
            missing = [k for k in ['ho_succ', 'ho_att'] if k not in field_map]
            self.log(f"[4G全程完好率] 缺少字段无法计算4G切换成功率: {missing}", "WARNING")

        # ========== 3. 计算4G E-RAB掉线率 ==========
        if all(k in field_map for k in ['ho_fail', 'erab_normal_rel', 'erab_rel', 'context_left', 'erab_succ', 'ho_inc']):
            ho_fail = df[field_map['ho_fail']].fillna(0).astype(float)
            erab_normal_rel = df[field_map['erab_normal_rel']].fillna(0).astype(float)
            erab_rel = df[field_map['erab_rel']].fillna(0).astype(float)
            context_left = df[field_map['context_left']].fillna(0).astype(float)
            erab_succ = df[field_map['erab_succ']].fillna(0).astype(float)
            ho_inc = df[field_map['ho_inc']].fillna(0).astype(float)

            numerator = ho_fail - erab_normal_rel + erab_rel
            denominator = context_left + erab_succ + ho_inc
            df['4G_E-RAB掉线率(%)'] = np.where(denominator > 0, numerator / denominator * 100, np.nan)
            df['4G_E-RAB掉线率(%)'] = df['4G_E-RAB掉线率(%)'].round(4)
            valid_count = df['4G_E-RAB掉线率(%)'].notna().sum()
            self.log(f"[4G全程完好率] 4G E-RAB掉线率计算完成, 有效数据: {valid_count}", "SUCCESS")
        else:
            missing = [k for k in ['ho_fail', 'erab_normal_rel', 'erab_rel', 'context_left', 'erab_succ', 'ho_inc'] if k not in field_map]
            self.log(f"[4G全程完好率] 缺少字段无法计算4G E-RAB掉线率: {missing}", "WARNING")

        # ========== 4. 计算4G全程完好率 ==========
        if '4G无线接通率(%)' in df.columns and '4G切换成功率(%)' in df.columns and '4G_E-RAB掉线率(%)' in df.columns:
            df['4G全程完好率(%)'] = (
                df['4G无线接通率(%)'] *
                df['4G切换成功率(%)'] *
                (100 - df['4G_E-RAB掉线率(%)'])
            ).round(4)
            valid_count = df['4G全程完好率(%)'].notna().sum()
            self.log(f"[4G全程完好率] 4G全程完好率计算完成, 有效数据: {valid_count}", "SUCCESS")
        else:
            self.log(f"[4G全程完好率] 无法计算4G全程完好率(缺少前置指标)", "WARNING")

        # ========== 5. 判断4G是否差小区 ==========
        if '4G全程完好率(%)' in df.columns:
            df['4G是否差小区'] = np.where(df['4G全程完好率(%)'] < 85, '是', '否')
            bad_count = (df['4G是否差小区'] == '是').sum()
            self.log(f"[4G全程完好率] 4G是否差小区计算完成, 差小区数量: {bad_count}", "SUCCESS")

        return df

    def _add_5g_wanchenglv_calc_columns(self, df):
        """添加5G全程完好率计算列

        计算规则：
        - SA无线接通率% = (RRC连接建立成功次数/RRC连接建立请求次数)
                        * (Flow建立成功数/Flow建立请求数)
                        * (NG接口UE相关逻辑信令连接建立成功次数/NG接口UE相关逻辑信令连接建立请求次数)
        - SA无线掉线率% = (gNB请求释放上下文数 - 正常的gNB请求释放上下文数)
                       / (初始上下文建立成功次数 + 遗留上下文个数 + 切换入成功次数 + RRC连接重建成功次数(非源侧小区))
        - SA切换成功率% = (gNB间NG切换出成功次数 + gNB间Xn切换出成功次数 + CU内DU间切换出执行成功次数 + CU内DU内切换出成功次数)
                        / (gNB间NG切换出准备请求次数 + gNB间Xn切换出准备请求次数 + CU内DU间切换出执行请求次数 + CU内DU内切换出执行请求次数)
        - 5G全程完好率 = SA无线接通率% * SA切换成功率% * (100 - SA无线掉线率%)
        - 5G是否差小区 = 全程完好率 < 85% 时为"是"，否则为"否"

        Args:
            df: 5G全程完好率报表的DataFrame

        Returns:
            DataFrame: 添加了计算列的DataFrame
        """
        import numpy as np

        # 字段名映射（英文 -> 通用名）
        field_map = {}
        for col in df.columns:
            col_lower = col.lower() if isinstance(col, str) else ''
            # RRC连接建立成功次数
            if col_lower in ('rrc_succconnestab', 'rrc连接建立成功次数'):
                field_map['rrc_succ'] = col
            # RRC连接建立请求次数
            elif col_lower in ('rrc_attconnestab', 'rrc连接建立请求次数'):
                field_map['rrc_att'] = col
            # Flow建立成功数
            elif col_lower in ('flow_nbrsuccestab', 'flow建立成功数'):
                field_map['flow_succ'] = col
            # Flow建立请求数
            elif col_lower in ('flow_nbrattestab', 'flow建立请求数'):
                field_map['flow_att'] = col
            # NG接口UE相关逻辑信令连接建立成功次数
            elif col_lower in ('ngsig_connestabsucc', 'ng接口ue相关逻辑信令连接建立成功次数'):
                field_map['ngsig_succ'] = col
            # NG接口UE相关逻辑信令连接建立请求次数
            elif col_lower in ('ngsig_connestabatt', 'ng接口ue相关逻辑信令连接建立请求次数'):
                field_map['ngsig_att'] = col
            # gNB请求释放上下文数
            elif col_lower in ('context_attrelgnb', 'gnb请求释放上下文数'):
                field_map['context_rel'] = col
            # 正常的gNB请求释放上下文数
            elif col_lower in ('context_attrelgnb_normal', '正常的gnb请求释放上下文数'):
                field_map['context_rel_normal'] = col
            # 初始上下文建立成功次数
            elif col_lower in ('context_succinitalsetup', '初始上下文建立成功次数'):
                field_map['context_init_succ'] = col
            # 遗留上下文个数
            elif col_lower in ('context_nbrleft', '遗留上下文个数'):
                field_map['context_left'] = col
            # 切换入成功次数
            elif col_lower in ('ho_succexecinc', '切换入成功次数'):
                field_map['ho_inc_succ'] = col
            # RRC连接重建成功次数(非源侧小区)
            elif col_lower in ('rrc_succconnreestab_nonsrccell', 'rrc连接重建成功次数(非源侧小区)'):
                field_map['rrc_reestab_succ'] = col
            # gNB间NG切换出成功次数
            elif col_lower in ('ho_succoutintercung', 'gnb间ng切换出成功次数'):
                field_map['ho_ng_succ'] = col
            # gNB间Xn切换出成功次数
            elif col_lower in ('ho_succoutintercuxn', 'gnb间xn切换出成功次数'):
                field_map['ho_xn_succ'] = col
            # CU内DU间切换出执行成功次数
            elif col_lower in ('ho_succoutintracuinterdu', 'cu内du间切换出执行成功次数'):
                field_map['ho_cu_du_succ'] = col
            # CU内DU内切换出成功次数
            elif col_lower in ('ho_succoutintradu', 'cu内du内切换出成功次数'):
                field_map['ho_cu_intra_succ'] = col
            # gNB间NG切换出准备请求次数
            elif col_lower in ('ho_attoutintercung', 'gnb间ng切换出准备请求次数'):
                field_map['ho_ng_att'] = col
            # gNB间Xn切换出准备请求次数
            elif col_lower in ('ho_attoutintercuxn', 'gnb间xn切换出准备请求次数'):
                field_map['ho_xn_att'] = col
            # CU内DU间切换出执行请求次数
            elif col_lower in ('ho_attoutintracuinterdu', 'cu内du间切换出执行请求次数'):
                field_map['ho_cu_du_att'] = col
            # CU内DU内切换出执行请求次数
            elif col_lower in ('ho_attoutcuintradu', 'cu内du内切换出执行请求次数'):
                field_map['ho_cu_intra_att'] = col

        self.log(f"[5G全程完好率] 字段映射: {list(field_map.keys())}", "INFO")

        # ========== 1. 计算SA无线接通率% ==========
        rrc_ok = all(k in field_map for k in ['rrc_succ', 'rrc_att'])
        flow_ok = all(k in field_map for k in ['flow_succ', 'flow_att'])
        ngsig_ok = all(k in field_map for k in ['ngsig_succ', 'ngsig_att'])

        if rrc_ok and flow_ok and ngsig_ok:
            rrc_succ = df[field_map['rrc_succ']].fillna(0).astype(float)
            rrc_att = df[field_map['rrc_att']].fillna(0).astype(float)
            flow_succ = df[field_map['flow_succ']].fillna(0).astype(float)
            flow_att = df[field_map['flow_att']].fillna(0).astype(float)
            ngsig_succ = df[field_map['ngsig_succ']].fillna(0).astype(float)
            ngsig_att = df[field_map['ngsig_att']].fillna(0).astype(float)

            rrc_rate = np.where(rrc_att > 0, rrc_succ / rrc_att, 0)
            flow_rate = np.where(flow_att > 0, flow_succ / flow_att, 0)
            ngsig_rate = np.where(ngsig_att > 0, ngsig_succ / ngsig_att, 0)

            df['SA无线接通率(%)'] = np.where(
                (rrc_att > 0) & (flow_att > 0) & (ngsig_att > 0),
                rrc_rate * flow_rate * ngsig_rate * 100,
                np.nan
            )
            df['SA无线接通率(%)'] = df['SA无线接通率(%)'].round(4)
            valid_count = df['SA无线接通率(%)'].notna().sum()
            self.log(f"[5G全程完好率] SA无线接通率计算完成, 有效数据: {valid_count}", "SUCCESS")
        else:
            missing = []
            if not rrc_ok:
                missing.extend([k for k in ['rrc_succ', 'rrc_att'] if k not in field_map])
            if not flow_ok:
                missing.extend([k for k in ['flow_succ', 'flow_att'] if k not in field_map])
            if not ngsig_ok:
                missing.extend([k for k in ['ngsig_succ', 'ngsig_att'] if k not in field_map])
            self.log(f"[5G全程完好率] 缺少字段无法计算SA无线接通率: {missing}", "WARNING")

        # ========== 2. 计算SA无线掉线率% ==========
        drop_ok = all(k in field_map for k in ['context_rel', 'context_rel_normal', 'context_init_succ',
                                                  'context_left', 'ho_inc_succ', 'rrc_reestab_succ'])
        if drop_ok:
            context_rel = df[field_map['context_rel']].fillna(0).astype(float)
            context_rel_normal = df[field_map['context_rel_normal']].fillna(0).astype(float)
            context_init_succ = df[field_map['context_init_succ']].fillna(0).astype(float)
            context_left = df[field_map['context_left']].fillna(0).astype(float)
            ho_inc_succ = df[field_map['ho_inc_succ']].fillna(0).astype(float)
            rrc_reestab_succ = df[field_map['rrc_reestab_succ']].fillna(0).astype(float)

            numerator = context_rel - context_rel_normal
            denominator = context_init_succ + context_left + ho_inc_succ + rrc_reestab_succ

            df['SA无线掉线率(%)'] = np.where(denominator > 0, numerator / denominator * 100, np.nan)
            df['SA无线掉线率(%)'] = df['SA无线掉线率(%)'].round(4)
            valid_count = df['SA无线掉线率(%)'].notna().sum()
            self.log(f"[5G全程完好率] SA无线掉线率计算完成, 有效数据: {valid_count}", "SUCCESS")
        else:
            missing = [k for k in ['context_rel', 'context_rel_normal', 'context_init_succ',
                                    'context_left', 'ho_inc_succ', 'rrc_reestab_succ'] if k not in field_map]
            self.log(f"[5G全程完好率] 缺少字段无法计算SA无线掉线率: {missing}", "WARNING")

        # ========== 3. 计算SA切换成功率% ==========
        ho_succ_ok = all(k in field_map for k in ['ho_ng_succ', 'ho_xn_succ', 'ho_cu_du_succ', 'ho_cu_intra_succ'])
        ho_att_ok = all(k in field_map for k in ['ho_ng_att', 'ho_xn_att', 'ho_cu_du_att', 'ho_cu_intra_att'])

        if ho_succ_ok and ho_att_ok:
            ho_ng_succ = df[field_map['ho_ng_succ']].fillna(0).astype(float)
            ho_xn_succ = df[field_map['ho_xn_succ']].fillna(0).astype(float)
            ho_cu_du_succ = df[field_map['ho_cu_du_succ']].fillna(0).astype(float)
            ho_cu_intra_succ = df[field_map['ho_cu_intra_succ']].fillna(0).astype(float)
            ho_ng_att = df[field_map['ho_ng_att']].fillna(0).astype(float)
            ho_xn_att = df[field_map['ho_xn_att']].fillna(0).astype(float)
            ho_cu_du_att = df[field_map['ho_cu_du_att']].fillna(0).astype(float)
            ho_cu_intra_att = df[field_map['ho_cu_intra_att']].fillna(0).astype(float)

            succ_total = ho_ng_succ + ho_xn_succ + ho_cu_du_succ + ho_cu_intra_succ
            att_total = ho_ng_att + ho_xn_att + ho_cu_du_att + ho_cu_intra_att

            df['SA切换成功率(%)'] = np.where(att_total > 0, succ_total / att_total * 100, np.nan)
            df['SA切换成功率(%)'] = df['SA切换成功率(%)'].round(4)
            valid_count = df['SA切换成功率(%)'].notna().sum()
            self.log(f"[5G全程完好率] SA切换成功率计算完成, 有效数据: {valid_count}", "SUCCESS")
        else:
            missing = []
            if not ho_succ_ok:
                missing.extend([k for k in ['ho_ng_succ', 'ho_xn_succ', 'ho_cu_du_succ', 'ho_cu_intra_succ'] if k not in field_map])
            if not ho_att_ok:
                missing.extend([k for k in ['ho_ng_att', 'ho_xn_att', 'ho_cu_du_att', 'ho_cu_intra_att'] if k not in field_map])
            self.log(f"[5G全程完好率] 缺少字段无法计算SA切换成功率: {missing}", "WARNING")

        # ========== 4. 计算5G全程完好率 ==========
        if 'SA无线接通率(%)' in df.columns and 'SA切换成功率(%)' in df.columns and 'SA无线掉线率(%)' in df.columns:
            df['5G全程完好率(%)'] = (
                df['SA无线接通率(%)'] *
                df['SA切换成功率(%)'] *
                (100 - df['SA无线掉线率(%)'])
            ).round(4)
            valid_count = df['5G全程完好率(%)'].notna().sum()
            self.log(f"[5G全程完好率] 5G全程完好率计算完成, 有效数据: {valid_count}", "SUCCESS")
        else:
            self.log(f"[5G全程完好率] 无法计算5G全程完好率(缺少前置指标)", "WARNING")

        # ========== 5. 判断5G是否差小区 ==========
        if '5G全程完好率(%)' in df.columns:
            df['5G是否差小区'] = np.where(df['5G全程完好率(%)'] < 85, '是', '否')
            bad_count = (df['5G是否差小区'] == '是').sum()
            self.log(f"[5G全程完好率] 5G是否差小区计算完成, 差小区数量: {bad_count}", "SUCCESS")

        return df

    def _on_query_complete(self):
        """查询完成回调"""
        self.is_querying = False
        self.extract_btn.config(state=tk.NORMAL, text="▶ 开始提取")
        self.stop_btn.config(state=tk.DISABLED)
        self.status_text.config(text="查询完成")
        self.status_dot.config(fg='#22c55e')  # 绿色
        self.update_progress(100, 100, "查询完成")
        self.log("所有查询完成！", "SUCCESS")

    def _on_query_failed(self):
        """查询失败回调"""
        self.is_querying = False
        self.extract_btn.config(state=tk.NORMAL, text="▶ 开始提取")
        self.stop_btn.config(state=tk.DISABLED)
        self.status_text.config(text="查询失败")
        self.status_dot.config(fg='#ef4444')  # 红色
        self.reset_progress()

    def log(self, message, level="INFO"):
        """输出日志"""
        self.log_text.config(state='normal')

        tag_map = {
            'INFO': 'INFO',
            'ERROR': 'ERROR',
            'WARNING': 'WARNING',
            'SUCCESS': 'SUCCESS'
        }

        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log_line = f"[{timestamp}] {message}\n"

        self.log_text.insert(tk.END, log_line, (tag_map.get(level, 'INFO'),))
        self.log_text.see(tk.END)
        self.log_text.config(state='disabled')

    def load_config(self):
        """加载配置"""
        self.log("NQI工具已就绪", "INFO")
        self.log(f"支持的数据表: {', '.join(TableConfig.get_table_names())}", "INFO")

    def run(self):
        """运行应用"""
        self.root.protocol("WM_DELETE_WINDOW", self._on_closing)
        self.root.mainloop()

    def _on_closing(self):
        """窗口关闭事件"""
        if hasattr(self, '_time_monitor'):
            self._time_monitor.stop()
        self.root.destroy()

    def _show_activate_window(self):
        """显示用户码激活窗口（新融合方案）"""
        # 获取本机机器码
        hw_info = get_hw_info()
        machine_code = generate_machine_code(hw_info)

        # 创建激活窗口
        activate_win = tk.Toplevel(self.root)
        activate_win.title("授权激活")
        activate_win.geometry("600x450")
        activate_win.resizable(False, False)

        # 设置窗口在主窗口中间
        self.root.update_idletasks()
        x = (self.root.winfo_width() - 600) // 2 + self.root.winfo_x()
        y = (self.root.winfo_height() - 450) // 2 + self.root.winfo_y()
        activate_win.geometry(f"600x450+{x}+{y}")

        activate_win.transient(self.root)
        activate_win.grab_set()

        # 顶部标题
        header = tk.Frame(activate_win, bg='#165DFF', height=50)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        tk.Label(header, text="🎫 授权激活",
                font=('Microsoft YaHei UI', 16, 'bold'),
                bg='#165DFF', fg='white').pack(pady=12)

        # 主内容
        content = tk.Frame(activate_win, bg='#f9fafb')
        content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # 本机信息卡片
        info_card = tk.Frame(content, bg='white')
        info_card.pack(fill=tk.X, pady=(0, 15))

        tk.Label(info_card, text="📋 本机信息",
                font=('Microsoft YaHei UI', 12, 'bold'),
                bg='white', fg='#374151', anchor='w').pack(padx=15, pady=(12, 5))

        machine_frame = tk.Frame(info_card, bg='white')
        machine_frame.pack(fill=tk.X, padx=15, pady=(0, 12))

        tk.Label(machine_frame, text="机器码：",
                font=('Microsoft YaHei UI', 9, 'bold'),
                bg='white', fg='#5f6368').pack(side=tk.LEFT)

        machine_code_label = tk.Label(machine_frame, text=machine_code,
                                    font=('Consolas', 8),
                                    bg='white', fg='#5f6368')
        machine_code_label.pack(side=tk.LEFT, padx=(5, 0))

        tk.Button(machine_frame, text="📋 复制",
                font=('Microsoft YaHei UI', 8),
                bg='#f0f2f5', fg='#202124', bd=1,
                cursor='arrow', relief='raised', padx=8, pady=2,
                command=lambda: self._copy_to_clipboard(activate_win, machine_code)).pack(side=tk.RIGHT)

        # 当前授权状态卡片
        status_card = tk.Frame(content, bg='white')
        status_card.pack(fill=tk.X, pady=(0, 15))

        tk.Label(status_card, text="📊 当前授权状态",
                font=('Microsoft YaHei UI', 12, 'bold'),
                bg='white', fg='#374151', anchor='w').pack(padx=15, pady=(12, 8))

        status_frame = tk.Frame(status_card, bg='white')
        status_frame.pack(fill=tk.X, padx=15, pady=(0, 12))

        # 检查当前授权状态
        info = get_user_code_info()
        if info:
            if info['days_left'] == -1:
                status_text = f"✅ 已授权（永久）"
                status_color = '#22c55e'
            elif info['days_left'] > 0:
                status_text = f"✅ 已授权（剩余 {info['days_left']} 天，到期 {info['expiry_date']}）"
                status_color = '#22c55e'
            else:
                status_text = f"❌ 授权已过期（{info['expiry_date']}）"
                status_color = '#ef4444'
        else:
            status_text = "❌ 未授权"
            status_color = '#ef4444'

        status_label = tk.Label(status_frame, text=status_text,
                               font=('Microsoft YaHei UI', 10),
                               bg='white', fg=status_color)
        status_label.pack(anchor='w')

        # 激活输入卡片
        input_card = tk.Frame(content, bg='white')
        input_card.pack(fill=tk.BOTH, expand=True)

        tk.Label(input_card, text="🔑 输入用户码",
                font=('Microsoft YaHei UI', 12, 'bold'),
                bg='white', fg='#374151', anchor='w').pack(padx=15, pady=(12, 5))

        tk.Label(input_card, text="请输入管理员提供的用户码：",
                font=('Microsoft YaHei UI', 9),
                bg='white', fg='#9ca3af', anchor='w').pack(padx=15, pady=(0, 8))

        serial_frame = tk.Frame(input_card, bg='white')
        serial_frame.pack(fill=tk.X, padx=15, pady=(0, 15))

        serial_entry = tk.Entry(serial_frame,
                              font=('Consolas', 10),
                              relief='flat', bg='#f8f9fa', bd=0)
        serial_entry.pack(fill=tk.X, ipady=8)

        tk.Label(serial_frame, text="用户码为 Base64 编码的字符串（由管理员生成）",
                font=('Microsoft YaHei UI', 8),
                bg='white', fg='#9ca3af').pack(anchor='w', pady=(4, 0))

        # 按钮
        btn_frame = tk.Frame(content, bg='#f9fafb')
        btn_frame.pack(fill=tk.X, pady=(15, 0))

        activate_btn = tk.Button(btn_frame, text="✅ 激活授权",
                 font=('Microsoft YaHei UI', 11, 'bold'),
                 bg='#22c55e', fg='white', bd=1,
                 cursor='hand2', relief='raised', padx=25, pady=8,
                 command=lambda: self._do_activate(serial_entry.get(), machine_code, activate_win))
        activate_btn.pack(side=tk.LEFT)

        # 注销按钮（如果已授权）
        if info:
            tk.Button(btn_frame, text="注销授权",
                     font=('Microsoft YaHei UI', 10),
                     bg='#fee2e2', fg='#dc2626', bd=1,
                     cursor='arrow', relief='raised', padx=18, pady=8,
                     command=lambda: self._do_deactivate(activate_win)).pack(side=tk.LEFT, padx=(10, 0))

        tk.Button(btn_frame, text="取消",
                 font=('Microsoft YaHei UI', 10),
                 bg='#f0f2f5', fg='#202124', bd=1,
                 cursor='arrow', relief='raised', padx=18, pady=8,
                 command=activate_win.destroy).pack(side=tk.RIGHT)

        serial_entry.focus()

        # 回车激活
        serial_entry.bind('<Return>', lambda e: activate_btn.invoke())

    def _do_activate(self, user_code, machine_code, window):
        """执行激活操作（新融合方案）

        流程：
        1. 解析用户码
        2. 验证机器码匹配
        3. 保存用户码
        """
        user_code = user_code.strip()
        if not user_code:
            messagebox.showwarning("提示", "请输入用户码")
            return

        # 导入解密函数进行验证
        from core.license import decrypt_user_code

        # 验证用户码
        success, expiry_timestamp, auth_machine_code = decrypt_user_code(user_code)

        if not success:
            messagebox.showerror("激活失败", "用户码格式错误或解密失败，请检查是否复制完整")
            return

        # 验证机器码匹配
        if auth_machine_code != machine_code:
            messagebox.showerror("激活失败", "用户码与本机机器码不匹配\n\n请确认您使用的是本机的用户码")
            return

        # 检查是否过期
        if expiry_timestamp != 0:
            from datetime import datetime as dt
            expiry_date = dt.fromtimestamp(expiry_timestamp).strftime('%Y-%m-%d')
            if expiry_timestamp < int(time.time()):
                messagebox.showerror("激活失败", f"用户码已过期（{expiry_date}）")
                return

        # 保存用户码
        if save_user_code(user_code):
            window.destroy()
            days_left = "永久" if expiry_timestamp == 0 else f"剩余 {(expiry_timestamp - int(time.time())) // 86400} 天"
            messagebox.showinfo("成功", f"授权激活成功！\n\n到期时间：{expiry_date if expiry_timestamp != 0 else '永久'}\n状态：{days_left}")
            # 重新加载授权信息
            self._reload_license()
        else:
            messagebox.showerror("错误", "保存用户码失败")

    def _do_deactivate(self, window):
        """注销授权"""
        if messagebox.askyesno("确认注销", "确定要注销当前授权吗？\n注销后需要重新激活才能使用。"):
            if delete_user_code():
                window.destroy()
                messagebox.showinfo("成功", "授权已注销")
                self._reload_license()
            else:
                messagebox.showerror("错误", "注销授权失败")

    def _reload_license(self):
        """重新加载授权信息"""
        # 重新验证授权
        hw_info = get_hw_info()
        machine_code = generate_machine_code(hw_info)

        valid, result = verify_with_user_code(machine_code)

        if valid:
            # 更新过期时间显示
            if result['days_left'] == -1:
                self.expiry_time = datetime(2099, 12, 31)  # 永久
            else:
                self.expiry_time = datetime.fromtimestamp(result['expiry_timestamp'])
            self._update_license_display()
            self.status_text.config(text="系统就绪")
            self.status_dot.config(fg='#22c55e')
            self.activate_btn.config(bg='#22c55e')  # 绿色表示已激活
        else:
            self.activate_btn.config(bg='#f59e0b')  # 橙色表示需要激活

    def _copy_to_clipboard(self, window, text):
        """复制到剪贴板"""
        window.clipboard_clear()
        window.clipboard_append(text)
        messagebox.showinfo("成功", "已复制到剪贴板")
