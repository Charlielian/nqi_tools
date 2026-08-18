# -*- coding: utf-8 -*-
"""
Excel样式管理器
提供统一的Excel样式定义和格式化功能
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math
import numbers
import pandas as pd


def normalize_excel_value(value):
    """将 Excel writer 不支持的缺失或非有限值转换为空单元格。"""
    try:
        missing = pd.isna(value)
        if isinstance(missing, bool) and missing:
            return None
        if hasattr(missing, 'item') and missing.ndim == 0 and bool(missing.item()):
            return None
        if isinstance(value, numbers.Number) and not isinstance(value, bool) and not math.isfinite(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def normalize_excel_rows(rows):
    return [[normalize_excel_value(value) for value in row] for row in rows]


class ExcelStyler:
    """Excel样式管理器"""

    # 预定义的样式配置
    STYLES = {
        'header_blue': {
            'color': '165DFF',
            'font_size': 11,
            'font_color': 'FFFFFF',
            'bold': True
        },
        'header_green': {
            'color': '22C55E',
            'font_size': 11,
            'font_color': 'FFFFFF',
            'bold': True
        },
        'header_orange': {
            'color': 'F59E0B',
            'font_size': 11,
            'font_color': 'FFFFFF',
            'bold': True
        },
        'header_red': {
            'color': 'EF4444',
            'font_size': 11,
            'font_color': 'FFFFFF',
            'bold': True
        },
    }

    # 边框样式
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 居中对齐
    CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
    WRAP_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

    @classmethod
    def get_header_style(cls, style_name='header_blue'):
        """获取表头样式

        Args:
            style_name: 样式名称

        Returns:
            tuple: (fill, font, alignment, border)
        """
        style = cls.STYLES.get(style_name, cls.STYLES['header_blue'])

        fill = PatternFill(
            start_color=style['color'],
            end_color=style['color'],
            fill_type='solid'
        )

        font = Font(
            bold=style['bold'],
            color=style['font_color'],
            size=style['font_size']
        )

        return fill, font, cls.WRAP_ALIGNMENT, cls.THIN_BORDER

    @classmethod
    def apply_header_style(cls, cell, style_name='header_blue'):
        """为单元格应用表头样式

        Args:
            cell: openpyxl Cell对象
            style_name: 样式名称
        """
        fill, font, alignment, border = cls.get_header_style(style_name)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border

    @classmethod
    def apply_cell_style(cls, cell, alignment_type='center'):
        """为单元格应用通用样式

        Args:
            cell: openpyxl Cell对象
            alignment_type: 对齐类型 ('center', 'left', 'right')
        """
        alignment_map = {
            'center': cls.CENTER_ALIGNMENT,
            'left': Alignment(horizontal='left', vertical='center'),
            'right': Alignment(horizontal='right', vertical='center')
        }
        cell.alignment = alignment_map.get(alignment_type, cls.CENTER_ALIGNMENT)
        cell.border = cls.THIN_BORDER

    @classmethod
    def auto_adjust_column_width(cls, worksheet, max_width=50, min_width=8):
        """自动调整列宽（openpyxl版）

        Args:
            worksheet: openpyxl Worksheet对象
            max_width: 最大列宽
            min_width: 最小列宽
        """
        for column in worksheet.columns:
            column_letter = get_column_letter(column[0].column)
            try:
                lengths = [len(str(cell.value)) for cell in column if cell.value is not None]
                max_length = max(lengths) if lengths else 8
            except (AttributeError, TypeError):
                max_length = 8

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    @classmethod
    def format_worksheet(cls, worksheet, header_style='header_blue',
                         auto_width=True, max_col_width=50):
        """格式化工作表（openpyxl版）

        Args:
            worksheet: openpyxl Worksheet对象
            header_style: 表头样式名称
            auto_width: 是否自动调整列宽
            max_col_width: 最大列宽
        """
        if worksheet.max_row < 1:
            return

        # 格式化表头
        header_fill, header_font, header_align, header_border = cls.get_header_style(header_style)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = header_border

        # 优化：使用 iter_rows 批量处理数据行
        if auto_width:
            # 预收集每列的最大长度
            col_max_lengths = {}
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    try:
                        col_idx = cell.column
                        val_len = len(str(cell.value)) if cell.value is not None else 0
                        if col_idx in col_max_lengths:
                            if val_len > col_max_lengths[col_idx]:
                                col_max_lengths[col_idx] = val_len
                        else:
                            col_max_lengths[col_idx] = val_len
                    except (AttributeError, TypeError):
                        pass

            # 批量应用样式和设置列宽
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = cls.CENTER_ALIGNMENT
                    cell.border = cls.THIN_BORDER

            # 批量设置列宽
            min_width = 8
            for col_idx, max_len in col_max_lengths.items():
                col_letter = get_column_letter(col_idx)
                adjusted_width = min(max(max_len + 2, min_width), max_col_width)
                worksheet.column_dimensions[col_letter].width = adjusted_width
        else:
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = cls.CENTER_ALIGNMENT
                    cell.border = cls.THIN_BORDER

    # ========== xlsxwriter 引擎支持（一次性写入+格式化，比openpyxl快3-5倍） ==========

    @classmethod
    def format_worksheet_xlsx(cls, workbook, worksheet, df, header_color='165DFF'):
        """使用xlsxwriter引擎格式化工作表（一次性写入样式）

        Args:
            workbook: xlsxwriter Workbook对象
            worksheet: xlsxwriter Worksheet对象
            df: pandas DataFrame
            header_color: 表头颜色（hex，不含#）
        """
        # 表头格式
        hdr_fmt = workbook.add_format({
            'bold': True,
            'font_color': '#FFFFFF',
            'bg_color': f'#{header_color}',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'text_wrap': True,
            'font_size': 11,
        })

        # 数据单元格格式
        cell_fmt = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
        })

        # 重写表头（用格式覆盖）
        headers = list(df.columns)
        for col_idx, col_name in enumerate(headers):
            worksheet.write(0, col_idx, col_name, hdr_fmt)

        # 批量写入数据行格式
        data_rows = normalize_excel_rows(df.values.tolist())
        for row_idx, row in enumerate(data_rows, start=1):
            worksheet.write_row(row_idx, 0, row, cell_fmt)

        # 自动列宽
        cls.auto_adjust_column_width_xlsx(workbook, worksheet, headers, df)

    @classmethod
    def auto_adjust_column_width_xlsx(cls, workbook, worksheet, headers, df, max_width=50, min_width=8):
        """自动调整列宽（xlsxwriter版 - 向量化计算）

        Args:
            workbook: xlsxwriter Workbook对象
            worksheet: xlsxwriter Worksheet对象
            headers: 列名列表
            df: pandas DataFrame
            max_width: 最大列宽
            min_width: 最小列宽
        """
        for col_idx, header in enumerate(headers):
            # 列宽 = max(表头长度, 数据列最大长度) + 2
            try:
                data_max = int(df.iloc[:, col_idx].astype(str).str.len().max())
            except (ValueError, TypeError):
                data_max = 0
            max_len = max(data_max, len(header))
            adjusted_width = min(max(max_len + 2, min_width), max_width)
            worksheet.set_column(col_idx, col_idx, adjusted_width)
